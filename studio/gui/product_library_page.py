# -*- coding: utf-8 -*-
"""
产品资料管理页。

把鼠标 / 键盘等外设按「品类 → 品牌 → 型号」统一归类管理：
- 基础数据通过服务端产品资料库接口（/api/product-library）一键同步
  （服务端统一持有旺店通 ERP 凭据并完成 ERP 拉取/归类，客户端不再直连旺店通）
- 左侧树状浏览 + 关键词搜索
- 右侧表单可对单条做手工归类 / 补充（品类、备注等）/ 删除

数据层见 utils/product_library_manager.py。
文案创作对接为后续步骤（manager.to_prompt_text 已预留）。
"""
from PySide6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QPushButton, QLineEdit, QTextEdit,
    QFrame, QWidget, QTreeWidget, QTreeWidgetItem, QMessageBox, QComboBox,
    QSplitter, QScrollArea, QFormLayout, QSizePolicy, QFileDialog,
)
from PySide6.QtCore import Qt, QThread, QTimer, Signal
from utils.base_worker import BaseWorker

from utils.logger_utils import log
from utils.product_library_manager import ProductLibraryManager, FIELDS, REQUIRED_FIELDS, WAREHOUSE_FIELDS


from utils.file_dialog_utils import pick_file, pick_save_file
from utils.gui_icons import mdi_button
def _get_server_url() -> str:
    """读取 ai_config.json 中的统一服务端地址（与 llm_proxy / compile_video_page 一致）。"""
    try:
        import json
        from config.paths import AI_CONFIG_FILE
        import os
        if os.path.isfile(AI_CONFIG_FILE):
            with open(AI_CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            url = (cfg.get("compute_server_url") or "").strip().rstrip("/")
            if url:
                return url
    except Exception:
        pass
    return ""


class StockSyncWorker(BaseWorker):
    """后台线程：通过服务端产品资料库接口同步产品数据。

    流程（所有存储逻辑均在服务端，按机器码隔离）：
      1. POST /api/product-library/clients/{machine_id}/sync        触发服务端 ERP 同步
      2. GET  /api/product-library/clients/{machine_id}/sync/status  轮询同步进度直至完成
      服务端内部已完成：ERP 拉取 → 库存映射 → 品类归类 → 持久化存储
      客户端只需在同步结束后刷新本地缓存。
    """
    phase = Signal(str)              # 阶段文字
    progress = Signal(int, int)      # fetched, total
    finished = Signal(int, int)      # added, updated

    def run(self):
        import time
        import requests
        from utils.http_client import http_get, http_post
        from utils.product_library_manager import _get_machine_id
        try:
            base = _get_server_url()
            if not base:
                self.error.emit("未配置服务端地址，请在系统设置中填写统一计算节点地址。")
                return
            machine_id = _get_machine_id()
            api = f"{base.rstrip('/')}/api/product-library/clients/{machine_id}"

            # 1. 触发服务端 ERP 同步（服务端内部完成 ERP 拉取 + 存储 + 品类归类）
            self.phase.emit("正在触发服务端 ERP 同步...")
            try:
                r = http_post(f"{api}/sync", timeout=10)
                if r.status_code == 409:
                    self.phase.emit("服务端同步进行中，等待完成...")
                elif r.status_code != 200:
                    self.error.emit(f"触发同步失败: HTTP {r.status_code} {r.text[:200]}")
                    return
            except requests.exceptions.RequestException as e:
                self.error.emit(f"无法连接服务端: {e}")
                return

            # 2. 轮询同步状态
            self.phase.emit("正在等待服务端同步完成...")
            while True:
                try:
                    sr = http_get(f"{api}/sync/status", timeout=10)
                    st = sr.json() if sr.status_code == 200 else {}
                except Exception:
                    st = {}
                if not st.get("running", False):
                    if st.get("error"):
                        self.error.emit(f"服务端同步出错: {st['error']}")
                        return
                    added = int(st.get("added", 0) or 0)
                    updated = int(st.get("updated", 0) or 0)
                    self.phase.emit(f"服务端同步完成（新增 {added}、更新 {updated}）")
                    self.finished.emit(added, updated)
                    return
                phase_text = st.get("phase", "") or "同步中..."
                fetched = int(st.get("fetched", 0) or 0)
                total = int(st.get("total", 0) or 0)
                self.phase.emit(phase_text)
                self.progress.emit(fetched, total)
                time.sleep(2)
        except Exception as e:
            self.error.emit(str(e))


from gui.base_page import BasePage



class SingleMineWorker(BaseWorker):
    """后台线程：按文档化接口挖掘单个产品并取回结果。

    流程：POST /mine（item_ids=[当前产品]）→ 轮询 /mine/status → GET /items/{item_id}。
    （旧实现调用的 /mine-single 在服务端不存在，返回 404。）

    与「一键成片」共享同一套数据源（服务端 product_items 表），
    挖掘完成后自动持久化，无需客户端再手动保存。
    """
    result_ready = Signal(dict)   # {ok, features, selling_points, persisted, item_id}

    def __init__(self, server_url, machine_id, item_id, category, brand,
                 model_name, spec_name, notes, llm_model):
        super().__init__()
        self.server_url = server_url
        self.machine_id = machine_id
        self.item_id = item_id
        self.category = category
        self.brand = brand
        self.model_name = model_name
        self.spec_name = spec_name
        self.notes = notes
        self.llm_model = llm_model

    def run(self):
        import time
        import requests
        from utils.http_client import http_get, http_post
        api = (f"{self.server_url.rstrip('/')}/api/product-library"
               f"/clients/{self.machine_id}")
        headers = {"X-Machine-ID": self.machine_id, "Content-Type": "application/json"}
        try:
            # 1. 触发挖掘：item_ids 非空 = 只挖当前选中的产品
            r = http_post(
                f"{api}/mine",
                json={"item_ids": [self.item_id], "model": self.llm_model},
                headers=headers, timeout=10)
            if r.status_code == 409:
                pass  # 服务端已有挖掘任务在跑，直接进入轮询
            elif r.status_code != 200:
                try:
                    err = r.json().get("detail", r.text[:200])
                except Exception:
                    err = r.text[:200]
                self.error.emit(f"触发挖掘失败 (HTTP {r.status_code}): {err}")
                return

            # 2. 轮询挖掘进度（最长 10 分钟）
            deadline = time.time() + 600
            while time.time() < deadline:
                try:
                    sr = http_get(f"{api}/mine/status", headers=headers, timeout=10)
                    st = sr.json() if sr.status_code == 200 else {}
                except Exception:
                    st = {}
                if not st.get("running", False):
                    if st.get("error"):
                        self.error.emit(f"服务端挖掘出错: {st['error']}")
                        return
                    break
                time.sleep(2)
            else:
                self.error.emit("服务端挖掘超时（10 分钟），请稍后在批量挖掘中查看。")
                return

            # 3. 拉取挖掘后的产品数据（features / selling_points 由服务端写入）。
            #    注意：GET /items/{id} 返回包裹结构 {"ok": true, "item": {...}}，
            #    需要取 item 层；并小幅重试以兼容服务端状态与落库的微小时序差。
            item = None
            for _attempt in range(8):
                g = http_get(f"{api}/items/{self.item_id}", headers=headers, timeout=10)
                if g.status_code == 200:
                    payload = g.json()
                    candidate = payload.get("item") if isinstance(payload, dict) else None
                    if not isinstance(candidate, dict):
                        candidate = payload if isinstance(payload, dict) else {}
                    item = candidate
                    if (str(candidate.get("features") or "").strip()
                            or str(candidate.get("selling_points") or "").strip()):
                        break
                time.sleep(1.5)
            if item is not None:
                self.result_ready.emit({
                    "ok": True,
                    "features": str(item.get("features") or ""),
                    "selling_points": str(item.get("selling_points") or ""),
                    "persisted": True,
                    "item_id": self.item_id,
                })
            else:
                self.error.emit("获取挖掘结果失败：服务端未返回产品数据。")
        except Exception as e:
            self.error.emit(f"请求服务端失败: {e}")


class BulkMineWorker(BaseWorker):
    """后台线程：触发服务端批量挖掘并轮询进度。

    服务端内部逐条调用 LLM + 写入 DB，客户端只需触发 + 轮询状态。
    """
    progress = Signal(int, int, str)   # done, total, phase_text
    finished = Signal(int, int)        # done_count, total

    def __init__(self, model):
        super().__init__()
        self.model = model
        self._should_stop = False

    def stop(self):
        self._should_stop = True

    def run(self):
        import time
        import requests
        from utils.http_client import http_get, http_post
        from utils.product_library_manager import _get_machine_id
        base = _get_server_url()
        if not base:
            self.error.emit("未配置服务端地址，请在系统设置中填写统一计算节点地址。")
            return
        machine_id = _get_machine_id()
        api = f"{base.rstrip('/')}/api/product-library/clients/{machine_id}"

        # 1. 触发服务端批量挖掘
        try:
            r = http_post(f"{api}/mine",
                          json={"item_ids": [], "model": self.model},
                          timeout=10)
            if r.status_code == 409:
                pass  # 已在运行中，直接轮询
            elif r.status_code != 200:
                self.error.emit(f"触发批量挖掘失败: HTTP {r.status_code} {r.text[:200]}")
                return
        except Exception as e:
            self.error.emit(f"无法连接服务端: {e}")
            return

        # 2. 轮询挖掘进度
        while not self._should_stop:
            try:
                sr = http_get(f"{api}/mine/status", timeout=10)
                st = sr.json() if sr.status_code == 200 else {}
            except Exception:
                st = {}
            if not st.get("running", False):
                if st.get("error"):
                    self.error.emit(f"服务端批量挖掘出错: {st['error']}")
                    return
                done = int(st.get("done", 0) or 0)
                total = int(st.get("total", 0) or 0)
                self.progress.emit(done, total, "挖掘完成")
                self.finished.emit(done, total)
                return
            done = int(st.get("done", 0) or 0)
            total = int(st.get("total", 0) or 0)
            self.progress.emit(done, total, f"服务端挖掘中 {done}/{total}")
            time.sleep(2)
        # 用户手动停止
        self.finished.emit(0, 0)

class ProductLibraryPage(BasePage):
    def __init__(self, parent_widget, main_window):
        super().__init__(parent_widget, main_window)
        self.manager = ProductLibraryManager()
        self.current_id = None          # 当前编辑条目 id；None = 新增模式
        self.inputs = {}                # field key -> 输入控件
        self.sync_worker = None
        self.bulk_mine_worker = None

    # ---------------- UI 构建 ----------------
    def setup(self):
        root = QVBoxLayout(self.parent_widget)
        root.setContentsMargins(40, 40, 40, 40)
        root.setSpacing(16)

        hdr = QHBoxLayout()
        heading = QLabel("📦 产品资料")
        heading.setObjectName("heading")
        hdr.addWidget(heading)

        subtitle = QLabel("基础数据从旺店通仓库同步，按 品类 → 品牌 → 型号 统一管理，供 AI 文案创作调用")
        subtitle.setObjectName("muted_text")
        subtitle.setWordWrap(True)
        subtitle.setMaximumWidth(1400)  # 一行显示，右侧留白避让资源监控
        hdr.addWidget(subtitle)
        hdr.addStretch()
        root.addLayout(hdr)

        # 顶部：仓库同步条
        sync_bar = QHBoxLayout()
        self.btn_sync = QPushButton("🔄 从仓库同步")
        self.btn_sync.setObjectName("primary_button")
        self.btn_sync.clicked.connect(self._on_sync)
        sync_bar.addWidget(self.btn_sync)
        self.btn_import_excel = mdi_button("导入表格", "folder")
        self.btn_import_excel.setObjectName("secondary_button")
        self.btn_import_excel.clicked.connect(self._on_import_excel)
        sync_bar.addWidget(self.btn_import_excel)
        self.btn_export_template = mdi_button("导出模板", "save")
        self.btn_export_template.setObjectName("secondary_button")
        self.btn_export_template.clicked.connect(self._on_export_template)
        sync_bar.addWidget(self.btn_export_template)
        self.btn_mine_all = QPushButton("⚡ 全量挖掘")
        self.btn_mine_all.setObjectName("secondary_button")
        self.btn_mine_all.setToolTip("批量为所有产品自动挖掘性能参数和核心卖点")
        self.btn_mine_all.clicked.connect(self._on_mine_all)
        sync_bar.addWidget(self.btn_mine_all)
        self.sync_status = QLabel("")
        self.sync_status.setObjectName("muted_text")
        sync_bar.addWidget(self.sync_status, 1)
        root.addLayout(sync_bar)

        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(self._build_left_panel())
        splitter.addWidget(self._build_right_panel())
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        root.addWidget(splitter, 1)

        self.refresh_tree()

    def _build_left_panel(self):
        panel = QFrame()
        panel.setObjectName("card")
        layout = QVBoxLayout(panel)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        title = QLabel("产品库")
        title.setObjectName("card_title")
        layout.addWidget(title)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索 品牌 / 型号 / 编码 / 条码 ...")
        # 搜索防抖：避免每次按键都重建 600+ 节点树并发一次服务端请求
        self._search_timer = QTimer(self.parent_widget)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(300)
        self._search_timer.timeout.connect(self.refresh_tree)
        self.search_input.textChanged.connect(self._search_timer.start)
        layout.addWidget(self.search_input)

        self.tree = QTreeWidget()
        self.tree.setHeaderHidden(True)
        self.tree.itemClicked.connect(self._on_tree_clicked)
        layout.addWidget(self.tree, 1)

        btn_new = QPushButton("➕ 新增型号（清空表单）")
        btn_new.setObjectName("secondary_button")
        btn_new.clicked.connect(self.clear_form)
        layout.addWidget(btn_new)

        return panel

    def _build_right_panel(self):
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(12)

        # Create vertical splitter
        v_splitter = QSplitter(Qt.Vertical)
        
        # --- Top frame: Basic Info ---
        top_panel = QFrame()
        top_panel.setObjectName("card")
        top_layout = QVBoxLayout(top_panel)
        top_layout.setContentsMargins(16, 16, 16, 16)
        top_layout.setSpacing(10)

        title_basic = QLabel("📋 产品基本资料")
        title_basic.setObjectName("card_title")
        top_layout.addWidget(title_basic)

        tip = QLabel("仓库同步的产品：仅「商品名称/型号、品类、备注」可改，其余为仓库只读数据；"
                     "所有修改仅保存在本地，不会回写仓库。")
        tip.setObjectName("muted_text")
        tip.setWordWrap(True)
        top_layout.addWidget(tip)

        form_container = QWidget()
        form_grid = QGridLayout(form_container)
        form_grid.setSpacing(6)
        form_grid.setContentsMargins(0, 0, 0, 0)
        
        # Set stretch factors for the 4 widget columns (1, 3, 5, 7) to distribute space evenly
        form_grid.setColumnStretch(1, 1)
        form_grid.setColumnStretch(3, 1)
        form_grid.setColumnStretch(5, 1)
        form_grid.setColumnStretch(7, 1)

        existing_cats = self.manager.categories()
        basic_fields = [(k, l, m) for k, l, m in FIELDS if k not in ("features", "selling_points")]
        
        # Exact positions for a compact 3-row layout (row, widget_col_index, widget_span)
        # widget_col_index is 0, 1, 2, or 3.
        field_positions = {
            "category":      (0, 0, 1),
            "brand":         (0, 1, 1),
            "model":         (0, 2, 1),
            "goods_no":      (0, 3, 1),
            
            "spec_no":       (1, 0, 1),
            "spec_name":     (1, 1, 1),
            "barcode":       (1, 2, 1),
            "stock_num":     (1, 3, 1),
            
            "available_num": (2, 0, 1),
            "warehouse":     (2, 1, 1),
            "notes":         (2, 2, 2), # Spans widget columns index 2 and 3 (grid columns 5, 6, 7)
        }

        for key, label, multiline in basic_fields:
            req = " *" if key in REQUIRED_FIELDS else ""
            if key == "category":
                w = QComboBox(); w.setEditable(True); w.addItems(existing_cats); w.setCurrentText("")
            elif key == "notes":
                # Convert notes to QLineEdit to fit perfectly on Row 3 alongside other single-line fields
                w = QLineEdit()
            else:
                w = QLineEdit()
            
            self.inputs[key] = w
            lbl = QLabel(label + req)
            lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            
            row, col_idx, col_span = field_positions[key]
            lbl_col = col_idx * 2
            widget_col = col_idx * 2 + 1
            
            form_grid.addWidget(lbl, row, lbl_col)
            if col_span > 1:
                # Spans multiple columns (specifically for notes: widget_col is 5, spans 3 columns to cover 5, 6, 7)
                form_grid.addWidget(w, row, widget_col, 1, col_span * 2 - 1)
            else:
                form_grid.addWidget(w, row, widget_col)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        scroll.setWidget(form_container)
        top_layout.addWidget(scroll, 1)

        v_splitter.addWidget(top_panel)

        # --- Bottom frame: AI Mining ---
        bottom_panel = QFrame()
        bottom_panel.setObjectName("card")
        bottom_layout = QVBoxLayout(bottom_panel)
        bottom_layout.setContentsMargins(16, 16, 16, 16)
        bottom_layout.setSpacing(10)

        bottom_title_bar = QHBoxLayout()
        title_ai = QLabel("✨ 智能挖掘 (性能参数 & 核心卖点)")
        title_ai.setObjectName("card_title")
        bottom_title_bar.addWidget(title_ai)
        
        self.btn_mine = QPushButton("🪄 智能挖掘")
        self.btn_mine.setObjectName("primary_button")
        self.btn_mine.clicked.connect(self._on_mine)
        bottom_title_bar.addWidget(self.btn_mine)
        bottom_layout.addLayout(bottom_title_bar)

        # Grid layout for AI parameters to support vertical stretching
        grid_ai = QGridLayout()
        grid_ai.setSpacing(10)
        grid_ai.setContentsMargins(0, 0, 0, 0)
        grid_ai.setColumnStretch(1, 1)
        grid_ai.setRowStretch(0, 1)
        grid_ai.setRowStretch(1, 1)

        # Create input widgets for AI fields
        lbl_features = QLabel("性能参数")
        lbl_features.setAlignment(Qt.AlignRight | Qt.AlignTop)
        self.inputs["features"] = QTextEdit()
        self.inputs["features"].setPlaceholderText("点击【智能挖掘】自动从大模型获取性能参数，或在此手动输入...")
        self.inputs["features"].setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.inputs["features"].setMinimumHeight(80)
        grid_ai.addWidget(lbl_features, 0, 0)
        grid_ai.addWidget(self.inputs["features"], 0, 1)

        lbl_selling = QLabel("核心卖点")
        lbl_selling.setAlignment(Qt.AlignRight | Qt.AlignTop)
        self.inputs["selling_points"] = QTextEdit()
        self.inputs["selling_points"].setPlaceholderText("点击【智能挖掘】自动从大模型获取核心卖点，或在此手动输入...")
        self.inputs["selling_points"].setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.inputs["selling_points"].setMinimumHeight(80)
        grid_ai.addWidget(lbl_selling, 1, 0)
        grid_ai.addWidget(self.inputs["selling_points"], 1, 1)

        bottom_layout.addLayout(grid_ai)
        v_splitter.addWidget(bottom_panel)

        # Set default splitter sizes (e.g. 180px for 3-row top basic info, rest 420px for AI Mining)
        v_splitter.setSizes([180, 420])
        container_layout.addWidget(v_splitter, 1)

        # --- Bottom Button Bar (outside splitter for consistency) ---
        btn_row = QHBoxLayout()
        self.btn_save = QPushButton("💾 保存（新增）")
        self.btn_save.setObjectName("primary_button")
        self.btn_save.clicked.connect(self._on_save)
        btn_row.addWidget(self.btn_save)

        self.btn_delete = QPushButton("🗑️ 删除")
        self.btn_delete.setObjectName("secondary_button")
        self.btn_delete.clicked.connect(self._on_delete)
        btn_row.addWidget(self.btn_delete)

        btn_clear = QPushButton("清空")
        btn_clear.setObjectName("secondary_button")
        btn_clear.clicked.connect(self.clear_form)
        btn_row.addWidget(btn_clear)

        btn_row.addStretch()
        self.status_label = QLabel("")
        self.status_label.setObjectName("muted_text")
        btn_row.addWidget(self.status_label)
        container_layout.addLayout(btn_row)

        return container

    def _on_mine(self):
        """智能挖掘：调用服务端 mine-single，挖掘即持久化。

        与「一键成片」共享同一套数据源（服务端 product_items 表），
        挖掘完成后自动持久化，无需手动保存。
        """
        # 1. 模型配置
        ai = self.ai_config
        model = ai.get("llm_model", "deepseek-v4-flash")
        if not model:
            QMessageBox.warning(self.parent_widget, "大模型未配置",
                                "请先在\u201cAI 设置 / 大模型配置\u201d中选择模型名称。")
            return

        # 2. 收集表单
        category = self._get_widget_value(self.inputs.get("category"))
        brand = self._get_widget_value(self.inputs.get("brand"))
        model_name = self._get_widget_value(self.inputs.get("model"))
        spec_name = self._get_widget_value(self.inputs.get("spec_name"))
        notes = self._get_widget_value(self.inputs.get("notes"))

        if not brand or not model_name:
            QMessageBox.warning(self.parent_widget, "信息不足",
                                "请确保产品\u201c品牌\u201d和\u201c型号/货品名称\u201d已填写！")
            return

        # 3. 服务端地址 + 机器码
        server_url = _get_server_url()
        if not server_url:
            QMessageBox.warning(self.parent_widget, "服务端未配置",
                                "请先在系统设置中配置统一计算节点地址。")
            return
        from utils.product_library_manager import _get_machine_id
        machine_id = _get_machine_id()

        # 4. 新增模式 → 先自动保存获得 item_id（持久化前提）
        item_id = self.current_id or ""
        if not item_id:
            data = self._collect_form()
            ok, msg, item = self.manager.add_item(data)
            if ok and item:
                self.current_id = item["id"]
                item_id = item["id"]
                self.btn_save.setText("\U0001f4be 保存修改")
                self._set_status("已自动保存新产品，正在挖掘...")
                self.refresh_tree()
            else:
                self._set_status(f"自动保存失败({msg})，挖掘结果将不会持久化...")

        # 5. 发起服务端挖掘
        self.btn_mine.setEnabled(False)
        self.btn_mine.setText("\u23f3 正在挖掘...")
        self._set_status("正在调用服务端 AI 挖掘（挖掘即持久化）...")

        self.mine_worker = SingleMineWorker(
            server_url, machine_id, item_id,
            category, brand, model_name, spec_name, notes, model
        )

        def _on_mine_done(data):
            self.btn_mine.setEnabled(True)
            self.btn_mine.setText("\U0001fa84 智能挖掘")
            features = data.get("features", "").strip()
            selling_points = data.get("selling_points", "").strip()
            persisted = data.get("persisted", False)

            if features or selling_points:
                self._set_widget_value(self.inputs["features"], features)
                self._set_widget_value(self.inputs["selling_points"], selling_points)
                # 刷新本地缓存，确保一键成片页能立即读取到最新数据
                self.manager.load()
                if persisted:
                    self._set_status("\u2705 AI 挖掘成功，已自动持久化（一键成片可直接使用）。")
                else:
                    self._set_status("\u2705 AI 挖掘成功！请点击\u201c保存修改\u201d以持久化数据。")
            else:
                self._set_status("AI 挖掘未返回有效结果，请重试或手动填写。")

        def _on_mine_err(err_msg):
            self.btn_mine.setEnabled(True)
            self.btn_mine.setText("\U0001fa84 智能挖掘")
            self._set_status(f"挖掘失败：{err_msg}")
            QMessageBox.critical(self.parent_widget, "挖掘失败",
                                 f"服务端 AI 挖掘请求失败：\n{err_msg}")

        self.mine_worker.result_ready.connect(_on_mine_done)
        self.mine_worker.error.connect(_on_mine_err)
        self.track_worker(self.mine_worker)
        self.mine_worker.start()


    # ---------------- 控件读写 ----------------
    def _get_widget_value(self, w):
        if isinstance(w, QComboBox):
            return w.currentText().strip()
        if isinstance(w, QTextEdit):
            return w.toPlainText().strip()
        return w.text().strip()

    def _set_widget_value(self, w, value):
        value = str(value or "")
        if isinstance(w, QComboBox):
            w.setCurrentText(value)
        elif isinstance(w, QTextEdit):
            w.setPlainText(value)
        else:
            w.setText(value)

    def _collect_form(self):
        return {key: self._get_widget_value(w) for key, w in self.inputs.items()}

    def _fill_form(self, data):
        for key, w in self.inputs.items():
            self._set_widget_value(w, data.get(key, ""))

    # 仓库只读字段：仓库同步的产品里这些不可手工改（仅 商品名称/型号、品类、备注 可改）。
    LOCKABLE_FIELDS = ("brand", "goods_no", "spec_no", "spec_name",
                       "barcode", "stock_num", "available_num", "warehouse")

    def _apply_field_locks(self, warehouse):
        """warehouse=True（仓库同步来的产品）时锁定只读字段；新增/手工录入时全部可改。"""
        for key in self.LOCKABLE_FIELDS:
            w = self.inputs.get(key)
            if w is None:
                continue
            if isinstance(w, QComboBox):
                w.setEnabled(not warehouse)
            else:
                w.setReadOnly(warehouse)

    # ---------------- 树 ----------------
    def refresh_tree(self):
        """刷新产品树（HTTP 放后台线程，避免服务端异常时卡界面）。"""
        from utils.thread_worker import TaskWorker as Worker
        if getattr(self, "_tree_worker", None) and self._tree_worker.isRunning():
            return
        keyword = self.search_input.text().strip() if hasattr(self, "search_input") else ""
        # search()/grouped() 会同步请求服务端，放后台线程执行
        w = Worker(self._fetch_tree_data, keyword)
        self._tree_worker = w
        w.finished.connect(self._on_tree_data_ready)
        w.error.connect(lambda e: log.error(f"刷新产品库失败: {e}"))
        w.start()

    def _fetch_tree_data(self, keyword):
        """后台线程：按关键词搜索或取 grouped 树（只碰 manager，不碰 UI）。"""
        if keyword:
            return keyword, list(self.manager.search(keyword))
        return keyword, self.manager.grouped()

    def _on_tree_data_ready(self, payload):
        keyword, tree = payload
        self.tree.clear()
        if keyword:
            items = tree or []
            for it in items:
                label = f"{it.get('brand','')} {it.get('model','')}".strip() or "(未命名)"
                node = QTreeWidgetItem([label])
                node.setData(0, Qt.UserRole, it.get("id"))
                self.tree.addTopLevelItem(node)
            self.tree.expandAll()
            return
        tree = tree or {}
        for cat in sorted(tree.keys()):
            cat_node = QTreeWidgetItem([f"📂 {cat}"])
            cat_node.setData(0, Qt.UserRole, None)
            self.tree.addTopLevelItem(cat_node)
            for brand in sorted(tree[cat].keys()):
                brand_node = QTreeWidgetItem([f"🏷️ {brand}"])
                brand_node.setData(0, Qt.UserRole, None)
                cat_node.addChild(brand_node)
                for it in sorted(tree[cat][brand], key=lambda x: x.get("model", "")):
                    label = it.get("model", "") or it.get("goods_no", "") or "(未命名)"
                    leaf = QTreeWidgetItem([label])
                    leaf.setData(0, Qt.UserRole, it.get("id"))
                    brand_node.addChild(leaf)
        self.tree.expandAll()

    def _on_tree_clicked(self, item, _col):
        item_id = item.data(0, Qt.UserRole)
        if not item_id:
            return
        record = self.manager.get(item_id)
        if not record:
            # 缓存未命中（服务端 /items 只返回前 50 条；/grouped 全量但可能滞后）：
            # 直接向服务端取该产品，保证任何树节点都能立即显示详情
            record = self._fetch_item_direct(item_id)
        if not record:
            self._set_status(f"未找到产品（id={item_id}），请稍后重试或先同步。")
            return
        self.current_id = item_id
        self._fill_form(record)
        is_warehouse = bool(record.get("goods_no") or record.get("spec_no"))
        self._apply_field_locks(is_warehouse)
        self.btn_save.setText("💾 保存修改")
        self._set_status(f"正在编辑：{record.get('brand','')} {record.get('model','')}"
                         + ("（仓库产品：仅可改 商品名称/品类/备注）" if is_warehouse else ""))

    def _fetch_item_direct(self, item_id):
        """缓存未命中时直接 GET /items/{item_id} 取单个产品（含 features/selling_points）。"""
        from utils.http_client import http_get
        from utils.product_library_manager import _get_machine_id
        base = _get_server_url()
        if not base:
            return None
        try:
            machine_id = _get_machine_id()
            url = (f"{base.rstrip('/')}/api/product-library/clients/"
                   f"{machine_id}/items/{item_id}")
            r = http_get(url, headers={"X-Machine-ID": machine_id}, timeout=10)
            if r.status_code == 200:
                payload = r.json()
                cand = payload.get("item") if isinstance(payload, dict) else None
                if isinstance(cand, dict):
                    return cand
        except Exception as e:
            log.warning(f"[产品库] 直接获取产品失败 {item_id}: {e}")
        return None

    # ---------------- 手动增删改 ----------------
    def clear_form(self):
        self.current_id = None
        for w in self.inputs.values():
            self._set_widget_value(w, "")
        self._apply_field_locks(False)  # 新增/手工录入：全部可改
        self.btn_save.setText("💾 保存（新增）")
        self._set_status("新增模式")

    def _on_save(self):
        data = self._collect_form()
        if self.current_id:
            ok, msg, _ = self.manager.update_item(self.current_id, data)
        else:
            ok, msg, item = self.manager.add_item(data)
            if ok:
                self.current_id = item["id"]
                self.btn_save.setText("💾 保存修改")
        self._set_status(msg)
        if ok:
            self._refresh_combo_choices()
            self.refresh_tree()
        else:
            QMessageBox.warning(self.parent_widget, "无法保存", msg)

    def _on_delete(self):
        if not self.current_id:
            self._set_status("当前为新增模式，无可删除条目。")
            return
        record = self.manager.get(self.current_id)
        name = f"{record.get('brand','')} {record.get('model','')}".strip() if record else ""
        reply = QMessageBox.question(
            self.parent_widget, "确认删除",
            f"确定删除「{name}」吗？此操作不可撤销。",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        if self.manager.remove_item(self.current_id):
            self._set_status("已删除。")
            self.clear_form()
            self.refresh_tree()
        else:
            self._set_status("删除失败。")

    def _refresh_combo_choices(self):
        for key, getter in (("category", self.manager.categories), ("brand", lambda: self.manager.brands())):
            w = self.inputs.get(key)
            if isinstance(w, QComboBox):
                cur = w.currentText()
                w.blockSignals(True)
                w.clear(); w.addItems(getter()); w.setCurrentText(cur)
                w.blockSignals(False)

    # ---------------- 仓库同步 ----------------
    def _on_sync(self):
        if self.sync_worker and self.sync_worker.isRunning():
            return
        self.btn_sync.setEnabled(False)
        self._set_sync_status("正在连接服务端产品资料库...")
        self.sync_worker = StockSyncWorker()
        self.sync_worker.phase.connect(self._set_sync_status)
        self.sync_worker.progress.connect(self._on_sync_progress)
        self.sync_worker.finished.connect(self._on_sync_done)
        self.sync_worker.error.connect(self._on_sync_err)
        self.sync_worker.start()

    def _on_sync_progress(self, fetched, total):
        self._set_sync_status(f"正在同步：{fetched}/{total or '?'}")

    def _on_sync_done(self, added, updated):
        """服务端已内部完成 ERP 同步+存储+归类，客户端只需刷新缓存。"""
        self.btn_sync.setEnabled(True)
        self.manager.load()  # 从服务端刷新本地缓存
        total_items = len(self.manager.items)
        self._set_sync_status(
            f"同步完成：新增 {added}、更新 {updated}，共 {total_items} 条。"
        )
        self._refresh_combo_choices()
        self.refresh_tree()

    def _on_sync_err(self, err):
        self.btn_sync.setEnabled(True)
        self._set_sync_status(f"同步失败：{err}")
        log.error(f"产品资料库同步失败: {err}")
        QMessageBox.critical(self.parent_widget, "同步失败",
                             f"从服务端同步产品资料失败：\n{err}\n\n"
                             "请检查：\n"
                             "1. 系统设置中「统一服务端地址」是否正确且服务端已启动；\n"
                             "2. 服务端是否已配置旺店通 ERP 凭据（由服务端统一持有，客户端无需本地 erp_config.json）。")

    # ---------------- 一键挖掘 ----------------
    def _on_mine_all(self):
        if self.bulk_mine_worker and self.bulk_mine_worker.isRunning():
            self.bulk_mine_worker.stop()
            self.btn_mine_all.setText("\u26a1 全量挖掘")
            self.btn_mine_all.setToolTip("批量为所有产品自动挖掘性能参数和核心卖点（跳过已有数据的产品）")
            self._set_sync_status("已请求停止\u2026")
            return
        ai = self.ai_config
        model = ai.get("llm_model", "deepseek-chat")
        if not model:
            QMessageBox.warning(self.parent_widget, "大模型未配置",
                                '请先在【AI 设置 / 大模型配置】中填写模型名称。')
            return
        server_url = _get_server_url()
        if not server_url:
            QMessageBox.warning(self.parent_widget, "服务端未配置",
                                "请先在系统设置中配置统一计算节点地址。")
            return
        total = len(self.manager.all_items())
        if total == 0:
            QMessageBox.information(self.parent_widget, "无产品", "产品库为空，请先同步仓库库存。")
            return
        already = sum(1 for it in self.manager.all_items()
                      if it.get("features", "").strip() and it.get("selling_points", "").strip())
        pending = total - already
        reply = QMessageBox.question(
            self.parent_widget, "一键挖掘确认",
            f"共 {total} 个产品，其中 {already} 个已有挖掘数据（将跳过），"
            f"即将为 {pending} 个产品调用服务端大模型挖掘性能参数与核心卖点。\n\n确认开始？",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.btn_mine_all.setText("\u23f9 停止挖掘")
        self.btn_mine_all.setToolTip("点击停止批量挖掘")
        self._set_sync_status(f"一键挖掘中\u2026（0/{pending}）")
        self.bulk_mine_worker = BulkMineWorker(model)
        self.bulk_mine_worker.progress.connect(self._on_mine_all_progress)
        self.bulk_mine_worker.finished.connect(self._on_mine_all_done)
        self.bulk_mine_worker.error.connect(self._on_mine_all_err)
        self.track_worker(self.bulk_mine_worker)
        self.bulk_mine_worker.start()

    def _on_mine_all_progress(self, done, total, msg):
        self._set_sync_status(f"一键挖掘中\u2026（{done}/{total}）{msg}")

    def _on_mine_all_done(self, done_count, total):
        self.btn_mine_all.setText("\u26a1 全量挖掘")
        self.btn_mine_all.setToolTip("批量为所有产品自动挖掘性能参数和核心卖点（跳过已有数据的产品）")
        self._set_sync_status(f"一键挖掘完成：处理 {done_count}/{total} 条。")
        self.manager.load()  # 刷新本地缓存
        self.refresh_tree()
        if self.current_id:
            record = self.manager.get(self.current_id)
            if record:
                self._fill_form(record)

    def _on_mine_all_err(self, err):
        self.btn_mine_all.setText("\u26a1 全量挖掘")
        self.btn_mine_all.setToolTip("批量为所有产品自动挖掘性能参数和核心卖点（跳过已有数据的产品）")
        self._set_sync_status(f"一键挖掘出错：{err}")


    # ---------------- Excel 导入导出 ----------------

    def _on_export_template(self):
        """导出 Excel 导入模板（当前数据格式）。"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from utils.product_library_manager import FIELDS
        path, _ = pick_save_file(
            self.parent_widget, "导出导入模板", "产品资料导入模板.xlsx",
            "Excel 文件 (*.xlsx)")
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "产品资料"
        keys = [f[0] for f in FIELDS]    # 内部字段名
        labels = [f[1] for f in FIELDS]  # 中文显示名
        # 表头（中文名）
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, label in enumerate(labels, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = header_fill
            cell.font = header_font
        # 填入一行示例数据
        for col, key in enumerate(keys, 1):
            sample_map = {"category": "鼠标", "brand": "罗技", "model": "G502",
                          "goods_no": "示例GD001", "spec_no": "示例SP001",
                          "spec_name": "示例规格", "barcode": "6901234567890"}
            ws.cell(row=2, column=col, value=sample_map.get(key, f"示例{labels[col-1]}"))
        # 设置列宽
        for col in range(1, len(keys) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        wb.save(path)
        QMessageBox.information(self.parent_widget, "提示",
                                f"模板已导出到：{path}\n\n请按表头格式填写数据，然后使用「导入表格」功能导入。")

    def _on_import_excel(self):
        """从 Excel 导入产品数据（通过服务端存储）。"""
        import openpyxl
        from utils.product_library_manager import FIELDS
        path, _ = pick_file(
            self.parent_widget, "选择 Excel 文件", "",
            "Excel 文件 (*.xlsx *.xls)")
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            field_keys = [f[0] for f in FIELDS]
            valid_items = []
            errors = []
            for i, row in enumerate(rows, 2):
                if all(v is None or str(v).strip() == "" for v in row):
                    continue  # 跳过空行
                item = {}
                for col, key in enumerate(field_keys):
                    val = row[col] if col < len(row) else None
                    item[key] = str(val).strip() if val is not None else ""
                if not item.get("category") or not item.get("brand"):
                    errors.append(f"第 {i} 行：分类和品牌不能为空")
                    continue
                valid_items.append(item)
            # 批量 upsert 到服务端
            added = updated = 0
            if valid_items:
                added, updated = self.manager.upsert_stocks(valid_items)
                self.refresh_tree()
            msg = f"导入完成：成功 {added + updated} 条（新增 {added}、更新 {updated}）" if valid_items else "未导入任何数据。"
            if errors:
                msg += f"\n失败 {len(errors)} 条：\n" + "\n".join(errors[:10])
                if len(errors) > 10:
                    msg += f"\n... 还有 {len(errors)-10} 条错误"
            QMessageBox.information(self.parent_widget, "导入结果", msg)
        except Exception as e:
            QMessageBox.critical(self.parent_widget, "导入失败", f"文件读取失败：{e}")

    # ---------------- 杂项 ----------------
    def _set_status(self, text):
        if hasattr(self, "status_label"):
            self.status_label.setText(text)

    def _set_sync_status(self, text):
        if hasattr(self, "sync_status"):
            self.sync_status.setText(text)
