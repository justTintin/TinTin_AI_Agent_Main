"""
「分镜脚本创作」页

流程：输入/接收视频文案 → 生成分镜脚本（竖向列表，每镜头含镜别/时长/音效/画面描述/旁白）
      → 每镜头可引用素材（本地/即梦/MG/联网） → 批量生成即梦镜头图 → 飞书同步
"""
import json
import os
import time
from datetime import datetime
from typing import Any

import requests.exceptions
from config.paths import CONFIG_INI_FILE, DREAMINA_OUTPUT_DIR, KNOWLEDGE_MEDIA_DIR
from gui._tab_compat import setup_tab_widget
from gui.ai_script_page import FeishuUploadWorker, LLMWorker
from gui.base_page import BasePage
from gui.elided_label import ElidedLabel
from gui.searchable_combo import SearchableComboBox
from gui.vector_search_page import VideoPreviewDialog, _ThumbWorker
from PySide6.QtCore import QRect, QSize, Qt, QTimer, Signal
from PySide6.QtGui import QColor, QCursor, QPainter, QPen, QPixmap
from PySide6.QtWidgets import (
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSpinBox,
    QSplitter,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from utils import material_client
from utils.base_worker import BaseWorker
from utils.dreamina_client import DreaminaClient
from utils.logger_utils import log
from utils.my_knowledge_manager import STYLIZATION_TYPE, MyKnowledgeManager

SHOT_TYPES = ["特写", "近景", "中景", "远景", "全景", "俯拍", "仰拍", "主观", "空镜"]

# ─────────────────────────────── Workers ────────────────────────────────────

class BatchShotImageWorker(BaseWorker):
    """逐镜头调用即梦 text2image 出图。"""
    phase = Signal(str)
    shot_done = Signal(int, str)
    finished = Signal(list)

    def __init__(self, shots, ratio, out_dir):
        super().__init__()
        self.shots = shots   # [(idx, prompt)]
        self.ratio = ratio
        self.out_dir = out_dir

    def do_work(self):
        client = DreaminaClient()
        if not client.is_installed():
            raise RuntimeError("未找到即梦 CLI，请先在『即梦生成』页安装二进制。")
        logged, _ = client.is_logged_in()
        if not logged:
            raise RuntimeError("即梦未登录，请先到『即梦生成』页登录。")
        total = len(self.shots)
        all_files = []
        for n, (idx, prompt) in enumerate(self.shots, 1):
            self.phase.emit(f"镜头 {n}/{total} 提交中…")
            _, info = client.text2image(prompt, ratio=self.ratio, poll=0)
            submit_id = info.get("submit_id", "")
            if not submit_id:
                self.phase.emit(f"镜头 {n}/{total} 提交失败，跳过")
                continue
            shot_dir = os.path.join(self.out_dir, f"shot_{idx:02d}")
            for _ in range(40):
                info2, files = client.query_result(submit_id, download_dir=shot_dir)
                if files:
                    all_files.extend(files)
                    self.shot_done.emit(idx, files[0])
                    break
                if info2.get("gen_status") == "fail":
                    self.phase.emit(f"镜头 {n}/{total} 生成失败")
                    break
                self.phase.emit(f"镜头 {n}/{total} 出图中…")
                time.sleep(6)
        self.finished.emit(all_files)

class SingleShotImageWorker(BaseWorker):
    """单镜头即梦 text2image（供「引用素材」对话框使用）。"""
    done = Signal(str)   # 生成图片的本地路径

    def __init__(self, prompt, ratio, out_dir):
        super().__init__()
        self.prompt = prompt
        self.ratio = ratio
        self.out_dir = out_dir

    def do_work(self):
        client = DreaminaClient()
        if not client.is_installed():
            raise RuntimeError("未找到即梦 CLI")
        _, info = client.text2image(self.prompt, ratio=self.ratio, poll=0)
        submit_id = info.get("submit_id", "")
        if not submit_id:
            raise RuntimeError("即梦提交失败")
        os.makedirs(self.out_dir, exist_ok=True)
        for _ in range(40):
            info2, files = client.query_result(submit_id, download_dir=self.out_dir)
            if files:
                self.done.emit(files[0])
                return
            if info2.get("gen_status") == "fail":
                raise RuntimeError("即梦生成失败")
            time.sleep(6)
        raise RuntimeError("即梦生成超时（>4 分钟）")

# ─────────────────────────── 相似度检索 Worker ──────────────────────────────

class _SimilarSearchWorker(BaseWorker):
    """按文案做 CLIP 向量相似度检索，返回按 score 排序的素材（已按 material 去重）。"""
    finished = Signal(list)   # [dict, ...]

    def __init__(self, query: str, top_k: int = 30,
                 filter_brand: str = "", filter_category: str = "",
                 filter_path_prefix: str = ""):
        super().__init__()
        self.query = query
        self.top_k = top_k
        self.filter_brand = filter_brand or None
        self.filter_category = filter_category or None
        self.filter_path_prefix = filter_path_prefix or None

    def do_work(self):
        """服务端 CLIP 向量检索（与素材检索页一致，POST /material/search）。"""
        params = {"query": self.query,
                  "limit": max(self.top_k * 3, self.top_k), "offset": 0}
        if self.filter_brand:
            params["brand"] = self.filter_brand
        if self.filter_category:
            params["category"] = self.filter_category
        try:
            data = material_client.search(params, timeout=20)
            if data is None:
                self.error.emit("服务端素材检索失败")
                return
            rows = data.get("results") or data.get("data") or []
        except requests.exceptions.RequestException as e:
            self.error.emit(f"服务端检索失败: {e}")
            return

        # 统一字段（服务端返回 id/filename/file_hash/media_type/brand/model/score…，
        # path 用服务端流式地址，便于预览/导出引用）
        def _norm(r):
            mid = str(r.get("id") or r.get("material_id") or "")
            return {
                "material_id": mid,
                "filename": r.get("filename", ""),
                "file_hash": r.get("file_hash", ""),
                "media_type": r.get("media_type", ""),
                "brand": r.get("brand", ""),
                "model": r.get("model", ""),
                "product": r.get("product", ""),
                "score": r.get("score", 0),
                "scene_desc_primary": r.get("scene_desc_primary", ""),
                "path": material_client.serve_url(mid) if mid else "",
            }

        # 按 material_id / path 去重，保留每个素材的最高分
        best: dict[str, Any] = {}
        for r in (_norm(x) for x in rows):
            key = r.get("material_id") or r.get("path")
            if not key:
                continue
            if key not in best or (r.get("score") or 0) > (best[key].get("score") or 0):
                best[key] = r
        merged = sorted(best.values(), key=lambda x: x.get("score", 0) or 0, reverse=True)  # noqa: E501
        self.finished.emit(merged[: self.top_k])

class _StockSearchWorker(BaseWorker):
    """服务端联网素材搜索（POST /material/stock_search，Pexels/Pixabay）。

    与 WebSearchWorker 的 DuckDuckGo 网页搜索不同，本 Worker
    专门用于「引用素材」对话框的联网素材 Tab：返回带缩略图/直链/时长/分辨率
    的可商用素材列表，供用户直接选中后以 URL 形式绑定镜头。
    """
    finished = Signal(str, list)  # provider, items

    def __init__(self, query, kind="all"):
        super().__init__()
        self.query = query.strip()
        self.kind = kind if kind in ("image", "video", "all") else "all"

    def do_work(self):
        if not self.query:
            self.finished.emit("", [])
            return
        try:
            data = material_client.stock_search(self.query, self.kind, timeout=20)
            if data is None:
                # stock_search 不可达：不抛出 error，UI 层提示用户
                self.finished.emit("", [])
                return
            provider = (data or {}).get("provider", "")
            items = (data or {}).get("items") or []
            # 统一字段供 UI 渲染 / 绑定
            normalized = []
            for it in items:
                url = it.get("url") or ""
                thumb = it.get("thumb") or url
                normalized.append({
                    "stock_id": str(it.get("id") or it.get("stock_id") or hash(url) & 0xFFFFFFFF),
                    "type": (it.get("type") or self.kind).lower(),
                    "provider": it.get("provider") or provider or "",
                    "author": it.get("author") or "",
                    "title": it.get("title") or it.get("filename") or "",
                    "thumb": thumb,
                    "url": url,
                    "duration_sec": it.get("duration_sec") or 0,
                    "width": it.get("width") or 0,
                    "height": it.get("height") or 0,
                })
            self.finished.emit(provider, normalized)
        except requests.exceptions.RequestException as e:
            self.error.emit(str(e))

class _AutoBindShotsWorker(BaseWorker):
    """为每个镜头按其画面描述做相似度检索，自动匹配最相似素材。"""
    progress = Signal(int, int)          # done, total
    finished = Signal(object)            # {shot_index: material_dict}（整数键，Signal(dict) 的 QVariantMap 会丢键）  # noqa: E501

    def __init__(self, shots: list, min_score: float = 0.0,
                 filter_brand: str = "", filter_category: str = "",
                 filter_path_prefix: str = ""):
        super().__init__()
        # shots: [(shot_index, query_text), ...]
        self.shots = shots
        self.min_score = min_score
        self.filter_brand = filter_brand or None
        self.filter_category = filter_category or None
        self.filter_path_prefix = filter_path_prefix or None

    def do_work(self):
        """逐镜头调用服务端 /material/search 向量检索，自动匹配最相似素材。"""
        result = {}
        total = len(self.shots)
        for n, (shot_idx, query) in enumerate(self.shots, 1):
            self.progress.emit(n, total)
            q = (query or "").strip()
            if not q:
                continue
            try:
                params = {"query": q, "limit": 5, "offset": 0}
                if self.filter_brand:
                    params["brand"] = self.filter_brand
                if self.filter_category:
                    params["category"] = self.filter_category
                data = material_client.search(params, timeout=20)
                rows = (data.get("results") or data.get("data") or []
                        if data is not None else [])
            except requests.exceptions.RequestException:
                rows = []
            if not rows:
                continue
            top = rows[0]
            score = float(top.get("score", 0) or 0)
            if score < self.min_score:
                continue
            mid = str(top.get("id") or top.get("material_id") or "")
            result[shot_idx] = {
                "type": "local",
                "path": material_client.serve_url(mid) if mid else "",
                "name": top.get("filename", ""),
                "hash": top.get("file_hash", ""),
                "score": score,
            }
        self.finished.emit(result)

# ─────────────────────────── 引用素材对话框 ─────────────────────────────────

class ShotMaterialDialog(QDialog):
    """每个镜头的「引用素材」弹窗，包含三个来源：本地素材/MG动画/联网素材。"""

    def __init__(self, shot_desc="", ratio="9:16", brand="", model="",
                 category="", shot_type="", extra_ctx="", style="", topic="",
                 main_window=None, parent=None):
        try:
            super().__init__(parent)
            log.info(f"[引用素材] ShotMaterialDialog.__init__ 开始: parent={parent}, parent_widget_type={type(parent).__name__ if parent else 'None'}")
            self.setWindowTitle("引用素材")
            self.resize(1300, 850)
            self.setMinimumSize(1100, 750)
            self.selected_material = None
            self.selected_materials = []     # 复选结果（本地素材）
            self._web_worker = None
            self._img_worker = None
            self._thumb_cache = {}           # 本地素材缩略图：{mid: QPixmap}
            self._web_thumb_cache = {}       # 联网素材缩略图：{stock_id: QPixmap}
            self._thumb_queue = []
            self._web_thumb_queue = []
            self._active_thumb = 0
            self._web_active_thumb = 0
            self._thumb_workers = []
            self._web_thumb_workers = []
            self._selected_mids = set()      # 本地素材角标选中 mid
            self._web_selected_ids = set()   # 联网素材角标选中 stock_id
            self._local_placeholder = self._make_local_placeholder()
            self._web_placeholder = self._make_web_placeholder()
            self._ratio = ratio
            self._main_window = main_window
            self.brand = (brand or "").strip()
            self.model = (model or "").strip()
            self.category = (category or "").strip()
            self.shot_type = (shot_type or "").strip()
            self._extra_ctx = (extra_ctx or "").strip()
            self.style = (style or "").strip()
            self._topic = (topic or "").strip()
            log.info("[引用素材] ShotMaterialDialog 初始化成员变量完成，开始 _setup")
            self._setup(shot_desc)
            log.info("[引用素材] ShotMaterialDialog.__init__ 完成")
        except Exception as e:
            log.exception(f"[引用素材] ShotMaterialDialog.__init__ 异常: {e}")
            raise

    def _setup(self, shot_desc):
        try:
            log.info(f"[引用素材] _setup 开始: shot_desc_len={len(shot_desc or '')}")
            layout = QVBoxLayout(self)
            layout.setSpacing(10)

            self._tab_bar, self._stack, self.tabs = setup_tab_widget(layout, 1)
            log.info("[引用素材] setup_tab_widget 完成")

            # 检索上下文：景别 + 品牌 + 型号 + 产品类型 + 风格 + 文案/选题兜底，帮助 CLIP 命中产品相关素材
            self._search_ctx = " ".join(
                x for x in (self.shot_type, self.brand, self.model, self.category,
                            self.style, self._extra_ctx) if x)

            # ── 脚本基本信息（所有 tab 可见，作为全部素材搜索的统一过滤上下文）──
            info_parts = []
            if self._topic:
                info_parts.append(f"选题：{self._topic[:40]}")
            prod_txt = " ".join(x for x in (self.brand, self.model, self.category) if x)
            if prod_txt:
                info_parts.append(f"产品：{prod_txt}")
            if self.style:
                info_parts.append(f"风格：{self.style}")
            if self.shot_type:
                info_parts.append(f"景别：{self.shot_type}")
            if self._ratio:
                info_parts.append(f"画幅：{self._ratio}")
            self._script_info = " ｜ ".join(info_parts)
            if self._script_info:
                layout.addWidget(self._muted(" 脚本信息（所有素材搜索自动带上）：" + self._script_info))

            # ── Tab 1: 本地素材 ──────────────────────────────────────────
            local_tab = QWidget()
            lt = QVBoxLayout(local_tab)
            lt.setSpacing(8)
            row = QHBoxLayout()
            self.local_input = QLineEdit()
            self.local_input.setPlaceholderText("自动带入景别/品牌/型号/产品类型 + 镜头文案检索素材库")
            base = shot_desc[:120] if shot_desc else ""
            self.local_input.setText((self._search_ctx + " " + base).strip() if self._search_ctx else base)  # noqa: E501
            self.local_input.returnPressed.connect(self._search_local)
            row.addWidget(self.local_input, 1)
            btn_local = QPushButton(" 相似度检索")
            btn_local.clicked.connect(self._search_local)
            row.addWidget(btn_local)
            lt.addLayout(row)
            # 素材库结果：缩略图网格（右上角角标勾选可多选，双击预览/播放，确认后绑定 Hash）
            self.local_list = QListWidget()
            self.local_list.setViewMode(QListWidget.IconMode)
            self.local_list.setIconSize(QSize(160, 160))
            self.local_list.setGridSize(QSize(185, 205))
            self.local_list.setResizeMode(QListWidget.Adjust)
            self.local_list.setMovement(QListWidget.Static)
            self.local_list.setSpacing(8)
            self.local_list.setUniformItemSizes(True)
            self.local_list.itemDoubleClicked.connect(self._preview_local_item)
            self.local_list.itemClicked.connect(self._on_item_clicked)
            lt.addWidget(self.local_list, 1)
            lt.addWidget(self._muted("勾选所需素材（可多选）；双击缩略图预览/播放；确认后素材 Hash 绑定到当前镜头。"))
            self._tab_bar.addTab(" 素材库")
            self._stack.addWidget(local_tab)

            # ── Tab 2: MG 动画 ───────────────────────────────────────────
            mg_tab = QWidget()
            mg = QVBoxLayout(mg_tab)
            mg.setSpacing(12)
            mg.addStretch()
            mg_lbl = QLabel("MG 动画素材适合用作开场/标题/卡点/数字增长等动态素材，\n"
                             "请在「MG 动画」页完成制作后，回到此处通过「素材库」选项卡选取。")
            mg_lbl.setAlignment(Qt.AlignCenter)
            mg_lbl.setWordWrap(True)
            mg.addWidget(mg_lbl)
            btn_mg_jump = QPushButton(" 跳转到 MG 动画页")
            btn_mg_jump.setObjectName("secondary_button")
            btn_mg_jump.clicked.connect(self._open_mg)
            mg.addWidget(btn_mg_jump, 0, Qt.AlignCenter)
            mg.addStretch()
            self._tab_bar.addTab(" MG动画")
            self._stack.addWidget(mg_tab)

            # ── Tab 3: 联网素材（服务端 /material/stock_search，Pexels/Pixabay 免版权）──
            web_tab = QWidget()
            wt = QVBoxLayout(web_tab)
            wt.setSpacing(8)
            web_row = QHBoxLayout()
            # 类型筛选：支持 image/video/all 三种
            self.web_kind_combo = QComboBox()
            self.web_kind_combo.addItem("全部", "all")
            self.web_kind_combo.addItem("图片", "image")
            self.web_kind_combo.addItem("视频", "video")
            self.web_kind_combo.setFixedWidth(80)
            self.web_kind_combo.currentIndexChanged.connect(self._search_web)
            web_row.addWidget(self.web_kind_combo)
            self.web_input = QLineEdit()
            self.web_input.setPlaceholderText("输入搜索词，联网查找免版权素材（服务端 Pexels/Pixabay，可多选确认绑定）")
            # 与本地素材一致：默认带入景别/品牌/型号/产品类型 + 镜头文案
            self.web_input.setText((self._search_ctx + " " + shot_desc)[:80] if (self._search_ctx or shot_desc) else "")  # noqa: E501
            self.web_input.returnPressed.connect(self._search_web)
            web_row.addWidget(self.web_input, 1)
            self.btn_web = QPushButton("联网搜索")
            self.btn_web.setObjectName("primary_button")
            self.btn_web.clicked.connect(self._search_web)
            web_row.addWidget(self.btn_web)
            wt.addLayout(web_row)
            self.web_pbar = QProgressBar()
            self.web_pbar.setRange(0, 0)
            self.web_pbar.setVisible(False)
            wt.addWidget(self.web_pbar)
            # 结果网格（与本地素材一致的相册式缩略图网格）
            self.web_list = QListWidget()
            self.web_list.setViewMode(QListWidget.IconMode)
            self.web_list.setIconSize(QSize(160, 160))
            self.web_list.setGridSize(QSize(185, 215))
            self.web_list.setResizeMode(QListWidget.Adjust)
            self.web_list.setMovement(QListWidget.Static)
            self.web_list.setSpacing(8)
            self.web_list.setUniformItemSizes(True)
            self.web_list.itemDoubleClicked.connect(self._preview_web_item)
            self.web_list.itemClicked.connect(self._on_web_item_clicked)
            wt.addWidget(self.web_list, 1)
            wt.addWidget(self._muted("勾选所需联网素材（可多选，确认后以 URL 形式绑定镜头；Pexels/Pixabay 免版权可商用）。"))
            self._tab_bar.addTab(" 联网素材")
            self._stack.addWidget(web_tab)

            # ── 底部按钮 ─────────────────────────────────────────────────
            btn_row = QHBoxLayout()
            btn_row.addStretch()
            self.btn_confirm = QPushButton("完成： 确认选择")
            self.btn_confirm.setObjectName("primary_button")
            self.btn_confirm.setEnabled(False)
            self.btn_confirm.clicked.connect(self._confirm)
            btn_row.addWidget(self.btn_confirm)
            btn_cancel = QPushButton("取消")
            btn_cancel.clicked.connect(self.reject)
            btn_row.addWidget(btn_cancel)
            layout.addLayout(btn_row)

            self._tab_bar.currentChanged.connect(self._stack.setCurrentIndex)
            self._tab_bar.currentChanged.connect(self._on_tab_changed)
            log.info("[引用素材] 开始自动加载本地素材搜索")
            try:
                self._search_local()   # auto-load local on open
            except Exception as e:
                log.exception(f"[引用素材] 自动加载本地素材搜索失败: {e}")
            # 联网素材也默认加载一次（带脚本上下文过滤）
            if self._search_ctx or self.web_input.text().strip():
                try:
                    QTimer(self).singleShot(0, self._search_web)
                except Exception as e:
                    log.exception(f"[引用素材] 自动加载联网素材搜索失败: {e}")
            log.info("[引用素材] _setup 完成")
        except Exception as e:
            log.exception(f"[引用素材] _setup 异常: {e}")
            raise

    @staticmethod
    def _muted(text):
        lbl = QLabel(text)
        lbl.setObjectName("muted_text")
        lbl.setWordWrap(True)
        return lbl

    def _with_ctx(self, user_text):
        """搜索必须携带脚本基本信息过滤；用户输入已含上下文时不重复拼接。"""
        t = (user_text or "").strip()
        if self._search_ctx and self._search_ctx not in t:
            return f"{self._search_ctx} {t}".strip()
        return t

    # ── 本地素材（CLIP 相似度检索）────────────────────────────────────
    def _search_local(self):
        query = self._with_ctx(self.local_input.text().strip())
        self.local_list.clear()
        self.btn_confirm.setEnabled(False)
        if not query:
            self.local_list.addItem(QListWidgetItem("请输入或保留镜头文案后点「相似度检索」"))
            return

        self.local_list.addItem(QListWidgetItem("正在做相似度检索…"))
        self._sim_worker = _SimilarSearchWorker(
            query, top_k=30,
            filter_brand=self.brand, filter_category=self.category)
        self._sim_worker.finished.connect(self._on_similar_results)
        self._sim_worker.error.connect(self._on_similar_error)
        self._sim_worker.start()

    def _on_similar_error(self, msg):
        self.local_list.clear()
        self.local_list.addItem(QListWidgetItem(f"检索失败: {msg}"))

    def _on_similar_results(self, found):
        self.local_list.blockSignals(True)
        self.local_list.clear()
        self._thumb_cache.clear()
        self._thumb_queue.clear()
        self._active_thumb = 0
        if not found:
            self.local_list.blockSignals(False)
            self.local_list.addItem(QListWidgetItem("未找到相似素材，请调整镜头文案或先入库分析素材"))
            return
        for f in found:
            mid = str(f.get("material_id") or "")
            b = f.get("brand", "") or ""
            m = f.get("model", "") or ""
            info = f"{b} {m}".strip()
            score = float(f.get("score", 0) or 0)
            fname = f.get("filename", "") or "?"
            label = f"{fname}\n[{score*100:.0f}%] {info}".strip()
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, {
                "path": f.get("path", ""),
                "filename": fname,
                "file_hash": f.get("file_hash", ""),
                "media_type": f.get("media_type", ""),
                "brand": b, "model": m,
                "product": f.get("product", ""),
                "score": score,
                "mid": mid,
            })
            self._apply_icon(mid, item)
            item.setToolTip(
                f"{f.get('path','')}\n相似度: {score*100:.1f}%\nHash: {f.get('file_hash','')}\n"  # noqa: E501
                f"画面: {f.get('scene_desc_primary','') or '—'}"
            )
            self.local_list.addItem(item)
            if mid:
                self._thumb_queue.append(mid)
        self.local_list.blockSignals(False)
        self._drain_thumbs()
        self._on_local_sel_changed()

    # ── 相册式多选（右上角角标，与素材检索页一致）──────────────────────
    @staticmethod
    def _draw_corner_badge(base_pm, checked):
        """在缩略图右上角绘制选择方框（未选=空框，选中=绿底白勾）。"""
        pm = QPixmap(base_pm)
        p = QPainter(pm)
        p.setRenderHint(QPainter.Antialiasing)
        box = 22
        x = pm.width() - box - 6
        y = 6
        rect = QRect(x, y, box, box)
        if checked:
            p.setBrush(QColor("#2ecc71"))
            p.setPen(QPen(QColor("white"), 1.6))
            p.drawRoundedRect(rect, 4, 4)
            pen = QPen(QColor("white"), 2.6)
            pen.setCapStyle(Qt.RoundCap)
            pen.setJoinStyle(Qt.RoundJoin)
            p.setPen(pen)
            p.drawLine(x + 5, y + box * 0.55, x + box * 0.42, y + box * 0.82)
            p.drawLine(x + box * 0.42, y + box * 0.82, x + box * 0.78, y + box * 0.22)
        else:
            p.setBrush(QColor(15, 15, 20, 190))
            p.setPen(QPen(QColor("#c3c6d2"), 1.4))
            p.drawRoundedRect(rect, 4, 4)
        p.end()
        return pm

    def _apply_icon(self, mid, lw_item):
        """按当前选择状态设置缩略图（右上角叠加选择角标）。"""
        mid = str(mid or "")
        base = self._thumb_cache[mid] if mid and mid in self._thumb_cache else self._local_placeholder
        if mid:
            base = self._draw_corner_badge(base, mid in self._selected_mids)
        lw_item.setIcon(base)

    def _on_item_clicked(self, item):
        """单击：仅点击图标右上角选择方框区域切换选中；其余位置仅记录（双击预览）。"""
        d = item.data(Qt.UserRole) or {}
        mid = str(d.get("mid") or "")
        if not mid:
            return
        vp_pos = self.local_list.viewport().mapFromGlobal(QCursor.pos())
        item_rect = self.local_list.visualItemRect(item)
        badge_zone = QRect(item_rect.right() - 42, item_rect.top() + 2, 40, 40)
        if not badge_zone.contains(vp_pos):
            return
        if mid in self._selected_mids:
            self._selected_mids.discard(mid)
        else:
            self._selected_mids.add(mid)
        self._apply_icon(mid, item)
        self._on_local_sel_changed()

    def _on_local_sel_changed(self):
        if self.tabs.currentIndex() != 0:
            return
        self.btn_confirm.setEnabled(len(self._checked_local_items()) > 0)

    def _checked_local_items(self):
        """返回当前角标选中的本地素材条目。"""
        out = []
        for i in range(self.local_list.count()):
            it = self.local_list.item(i)
            d = it.data(Qt.UserRole)
            if d and str(d.get("mid") or "") in self._selected_mids:
                out.append(it)
        return out

    # ── 缩略图异步加载（并发节流，与素材检索一致）──────────────────
    def _drain_thumbs(self):
        while self._active_thumb < 6 and self._thumb_queue:
            mid = self._thumb_queue.pop(0)
            if mid in self._thumb_cache:
                continue
            self._active_thumb += 1
            w = _ThumbWorker(mid)
            self._thumb_workers.append(w)
            # finished(str, bytes) 直接接 _on_thumb_ready；不可用 lambda 包裹，否则图片数据丢失
            w.finished.connect(self._on_thumb_ready)
            w.error.connect(lambda _m, _mid=mid: self._on_thumb_ready(_mid))
            w.start()

    def _on_thumb_ready(self, mid, data=b""):
        self._active_thumb = max(0, self._active_thumb - 1)
        if data:
            pm = QPixmap()
            if pm.loadFromData(data) and not pm.isNull():
                self._thumb_cache[mid] = pm
                for i in range(self.local_list.count()):
                    it = self.local_list.item(i)
                    d = it.data(Qt.UserRole)
                    if d and str(d.get("mid") or "") == str(mid):
                        self._apply_icon(mid, it)
                        break
        self._drain_thumbs()

    def _preview_local_item(self, item):
        """双击：图片大图预览 / 视频播放。"""
        d = item.data(Qt.UserRole)
        if not d or not d.get("path"):
            return
        mid = str(d.get("mid") or "")
        mtype = (d.get("media_type") or "").lower()
        if mtype == "video":
            VideoPreviewDialog(d["path"], d.get("filename", ""), self).exec()
        else:
            dlg = QDialog(self)
            dlg.setWindowTitle(f"预览 - {d.get('filename', '')}")
            dlg.setMinimumSize(480, 480)
            lay = QVBoxLayout(dlg)
            lbl = QLabel("加载中…")
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("background:#000;color:#888;")
            lay.addWidget(lbl, 1)

            def on_loaded(_mid, data):
                pm = QPixmap()
                if data and pm.loadFromData(data) and not pm.isNull():
                    lbl.setPixmap(pm.scaled(lbl.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))  # noqa: E501

            if mid:
                w = _ThumbWorker(mid)
                self._thumb_workers.append(w)
                w.finished.connect(on_loaded)
                w.start()
            dlg.exec()

    @staticmethod
    def _make_local_placeholder():
        pm = QPixmap(160, 160)
        pm.fill(QColor("#243b55"))
        p = QPainter(pm)
        p.setPen(QColor("#888"))
        p.drawText(pm.rect(), Qt.AlignCenter, "…")
        p.end()
        return pm

    @staticmethod
    def _make_web_placeholder():
        pm = QPixmap(160, 160)
        pm.fill(QColor("#243b55"))
        p = QPainter(pm)
        p.setPen(QColor("#888"))
        f = p.font()
        f.setPointSize(11)
        p.setFont(f)
        p.drawText(pm.rect(), Qt.AlignCenter, "免版权素材\n加载中…")
        p.end()
        return pm

    # ── 即梦生成 ─────────────────────────────────────────────────────
    def _dreamina_gen(self):
        prompt = self.dreamina_input.toPlainText().strip()
        if not prompt:
            QMessageBox.warning(self, "提示词为空", "请填写即梦生图提示词。")
            return
        # 强制带上脚本基本信息（产品/风格/景别）作为生成约束
        if self._search_ctx and self._search_ctx not in prompt:
            prompt = f"{self._search_ctx}，{prompt}"
        self.btn_dreamina_gen.setEnabled(False)
        self.dreamina_pbar.setVisible(True)
        self.dreamina_thumb.setText("生成中，请稍候（约 1-4 分钟）...")
        self._dreamina_file = None
        self.btn_dreamina_select.setEnabled(False)

        out_dir = os.path.join(DREAMINA_OUTPUT_DIR, "shot_ref_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
        self._img_worker = SingleShotImageWorker(prompt, self._ratio, out_dir)

        def on_done(path):
            self.btn_dreamina_gen.setEnabled(True)
            self.dreamina_pbar.setVisible(False)
            self._dreamina_file = path
            pm = QPixmap(path)
            if not pm.isNull():
                self.dreamina_thumb.setPixmap(
                    pm.scaled(self.dreamina_thumb.width(), 200, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            self.btn_dreamina_select.setEnabled(True)

        def on_err(msg):
            self.btn_dreamina_gen.setEnabled(True)
            self.dreamina_pbar.setVisible(False)
            self.dreamina_thumb.setText(f"生成失败：{msg}")

        self._img_worker.done.connect(on_done)
        self._img_worker.error.connect(on_err)
        self._img_worker.start()

    def _select_dreamina(self):
        if self._dreamina_file:
            self.selected_material = {
                "type": "dreamina",
                "path": self._dreamina_file,
                "name": os.path.basename(self._dreamina_file),
            }
            self.accept()

    # ── MG 动画 ──────────────────────────────────────────────────────
    def _open_mg(self):
        try:
            if self._main_window:
                # MG动画在素材生成页（index 31）的内部 Tab2
                self._main_window.switch_page(31)
                if hasattr(self._main_window, "switch_dreamina_tab"):
                    self._main_window.switch_dreamina_tab(2)
        except Exception:
            pass
        self.reject()

    # ── 联网搜索（服务端 /material/stock_search，缩略图网格）─────────
    def _search_web(self):
        query = self._with_ctx(self.web_input.text().strip())
        if not query:
            return
        kind = self.web_kind_combo.currentData() if hasattr(self, "web_kind_combo") else "all"
        self.btn_web.setEnabled(False)
        self.web_pbar.setVisible(True)
        self.web_list.clear()
        self._web_thumb_queue.clear()
        self._web_thumb_cache.clear()
        self._web_selected_ids.clear()
        self.web_list.addItem(QListWidgetItem("正在联网搜索，请稍候（Pexels/Pixabay）…"))
        self._web_worker = _StockSearchWorker(query, kind)
        self._web_worker.finished.connect(self._on_stock_results)
        self._web_worker.error.connect(lambda m: (
            self.btn_web.setEnabled(True),
            self.web_pbar.setVisible(False),
            self.web_list.clear(),
            self.web_list.addItem(QListWidgetItem(f"联网搜索失败: {m}")),
        ))
        self._web_worker.start()

    def _on_stock_results(self, provider, items):
        self.btn_web.setEnabled(True)
        self.web_pbar.setVisible(False)
        self.web_list.blockSignals(True)
        self.web_list.clear()
        if not items:
            self.web_list.blockSignals(False)
            self.web_list.addItem(QListWidgetItem(
                f"未找到相关在线素材（来源: {provider or '服务端'}），可换关键词重试。"))
            self._on_web_sel_changed()
            return
        for it in items:
            sid = str(it.get("stock_id") or "")
            stype = (it.get("type") or "").lower()
            meta = []
            if it.get("duration_sec"):
                meta.append(f"{it['duration_sec']}s")
            if it.get("width") and it.get("height"):
                meta.append(f"{it['width']}x{it['height']}")
            if it.get("author"):
                meta.append(f"© {it['author']}")
            title = (it.get("title") or "").strip() or (it.get("provider", "") or "免版权素材")
            label = title if len(title) <= 16 else title[:15] + "…"
            if meta:
                label += "\n" + " ".join(meta)
            lw = QListWidgetItem(label)
            lw.setData(Qt.UserRole, it)
            lw.setForeground(QColor("#d1d5db"))
            self._apply_web_icon(sid, lw, stype)
            if it.get("url"):
                lw.setToolTip(
                    f"{title}\n来源: {it.get('provider') or 'Pexels/Pixabay'}\n作者: {it.get('author') or '—'}\n"
                    f"尺寸: {(it.get('width') or '—')}x{(it.get('height') or '—')}\n直链: {it['url']}"
                )
            self.web_list.addItem(lw)
            thumb = it.get("thumb") or ""
            if thumb and sid and sid not in self._web_thumb_cache:
                self._web_thumb_queue.append((sid, thumb))
        self.web_list.blockSignals(False)
        self._drain_web_thumbs()
        self._on_web_sel_changed()

    # ── 联网素材缩略图加载（URL 下载 → QPixmap，并发节流）─────────────
    def _apply_web_icon(self, sid, lw_item, mtype):
        sid = str(sid or "")
        if sid and sid in self._web_thumb_cache:
            base = self._web_thumb_cache[sid]
        else:
            base = self._pm_web_stock_video if mtype == "video" else self._web_placeholder
        if sid:
            base = self._draw_corner_badge(base, sid in self._web_selected_ids)
        lw_item.setIcon(base)

    @property
    def _pm_web_stock_video(self):
        """联网视频类型占位缩略图：深蓝色 + 播放角标。"""
        if not hasattr(self, "__pm_web_video"):
            pm = QPixmap(160, 160)
            pm.fill(QColor("#2b2d42"))
            p = QPainter(pm)
            p.setPen(QColor("#9aa0a6"))
            f = p.font()
            f.setPointSize(11)
            p.setFont(f)
            p.drawText(pm.rect(), Qt.AlignCenter, "▶ 联网视频")
            p.end()
            self.__pm_web_video = pm
        return self.__pm_web_video

    def _drain_web_thumbs(self):
        from utils.http_client import http_get

        def _load(sid, url):
            """下载 URL 缩略图并缓存，更新图标。"""
            try:
                resp = http_get(url, timeout=15)
                data = resp.content if resp.status_code == 200 and resp.content else b""
            except Exception:  # URL 缩略图下载涉及 HTTP 请求
                data = b""
            self._web_active_thumb = max(0, self._web_active_thumb - 1)
            if data:
                pm = QPixmap()
                if pm.loadFromData(data) and not pm.isNull():
                    sc = pm.scaled(160, 160, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
                    # 居中裁剪到正方形
                    cropped = QPixmap(160, 160)
                    cropped.fill(QColor("#16161f"))
                    p = QPainter(cropped)
                    x = (160 - sc.width()) // 2
                    y = (160 - sc.height()) // 2
                    p.drawPixmap(x, y, sc)
                    p.end()
                    self._web_thumb_cache[sid] = cropped
                    # 刷新列表项图标
                    for i in range(self.web_list.count()):
                        it = self.web_list.item(i)
                        d = it.data(Qt.UserRole)
                        if d and str(d.get("stock_id") or "") == sid:
                            self._apply_web_icon(sid, it, (d.get("type") or "").lower())
                            break
            QTimer(self).singleShot(0, self._drain_web_thumbs)

        while self._web_active_thumb < 6 and self._web_thumb_queue:
            sid, url = self._web_thumb_queue.pop(0)
            if sid in self._web_thumb_cache:
                continue
            self._web_active_thumb += 1
            import threading as _th

            t = _th.Thread(target=_load, args=(sid, url), daemon=True)
            t.start()

    # ── 联网素材选择 / 预览 ──────────────────────────────────────────
    def _on_web_item_clicked(self, item):
        """单击：只有点击右上角选择方框区域才切换选中。"""
        d = item.data(Qt.UserRole) or {}
        sid = str(d.get("stock_id") or "")
        if not sid:
            return
        vp_pos = self.web_list.viewport().mapFromGlobal(QCursor.pos())
        item_rect = self.web_list.visualItemRect(item)
        badge_zone = QRect(item_rect.right() - 42, item_rect.top() + 2, 40, 40)
        if not badge_zone.contains(vp_pos):
            return
        if sid in self._web_selected_ids:
            self._web_selected_ids.discard(sid)
        else:
            self._web_selected_ids.add(sid)
        self._apply_web_icon(sid, item, (d.get("type") or "").lower())
        self._on_web_sel_changed()

    def _on_web_sel_changed(self):
        if self._tab_bar.currentIndex() != 2:
            return
        count = sum(1 for i in range(self.web_list.count())
                    if str((self.web_list.item(i).data(Qt.UserRole) or {}).get("stock_id") or "") in self._web_selected_ids)
        self.btn_confirm.setEnabled(count > 0)

    def _preview_web_item(self, item):
        """双击联网素材：视频预览 / 图片大图预览。"""
        d = item.data(Qt.UserRole) or {}
        url = d.get("url") or ""
        if not url:
            return
        title = d.get("title") or d.get("provider") or "免版权素材"
        stock_id = str(d.get("stock_id") or "")
        media_type = (d.get("type") or "").lower()
        if media_type == "video":
            VideoPreviewDialog(url, title, stock_id, "video", self).exec()
            return
        dlg = QDialog(self)
        dlg.setWindowTitle(f"联网预览 - {title}")
        dlg.setMinimumSize(900, 560)
        dlg.resize(1080, 680)
        v = QVBoxLayout(dlg)
        lbl = QLabel("加载中…")
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet("background:#000;color:#888;")
        v.addWidget(lbl, 1)
        info = QLabel(f"来源: {d.get('provider') or 'Pexels/Pixabay'} | 作者: {d.get('author') or '—'} | 直链: {url}")
        info.setObjectName("muted_text")
        info.setWordWrap(True)
        v.addWidget(info)

        def on_done(path):
            pm = QPixmap(path)
            if not pm.isNull():
                lbl.setPixmap(pm.scaled(lbl.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation))
            else:
                lbl.setText("图片加载失败")

        import threading as _th

        def _download():
            try:
                from utils.http_client import http_get
                r = http_get(url, timeout=30)
                if r.status_code == 200 and r.content:
                    import hashlib
                    tmp_dir = os.path.join(DREAMINA_OUTPUT_DIR, "_web_preview")
                    os.makedirs(tmp_dir, exist_ok=True)
                    ext = os.path.splitext(url.split("?")[0])[1].lower()[:4]
                    ext = ext if ext in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp") else ".jpg"
                    fpath = os.path.join(tmp_dir, hashlib.md5(url.encode("utf-8")).hexdigest() + ext)
                    with open(fpath, "wb") as f:
                        f.write(r.content)
                    QTimer(dlg).singleShot(0, lambda: on_done(fpath))
                    return
            except Exception:  # 下载联网图片预览涉及 HTTP/文件操作
                pass
            QTimer(dlg).singleShot(0, lambda: lbl.setText("图片加载失败"))

        t = _th.Thread(target=_download, daemon=True)
        t.start()
        dlg.exec()

    # ── Tab 切换时更新确认按钮 ───────────────────────────────────────
    def _on_tab_changed(self, idx):
        if idx == 0:
            self._on_local_sel_changed()
        elif idx == 1:
            self.btn_confirm.setEnabled(self._dreamina_file is not None)
        elif idx == 2:
            self._on_web_sel_changed()
        else:
            self.btn_confirm.setEnabled(False)

    # ── 确认 ─────────────────────────────────────────────────────────
    def _confirm(self):
        tab = self._tab_bar.currentIndex()
        if tab == 0:
            items = self._checked_local_items()
            if items:
                mats = []
                for it in items:
                    d = it.data(Qt.UserRole)
                    if not d:
                        continue
                    mats.append({
                        "type": "local",
                        "path": d.get("path", ""),
                        "name": d.get("filename") or d.get("name", ""),
                        "hash": d.get("file_hash", ""),
                        "mid": d.get("mid", ""),
                    })
                if mats:
                    self.selected_materials = mats
                    self.selected_material = mats[0]
                    self.accept()
        elif tab == 1:
            self._select_dreamina()
        elif tab == 2:
            # 联网素材：按勾选的角标绑定多个，素材以 URL/类型保存
            mats = []
            for i in range(self.web_list.count()):
                it = self.web_list.item(i)
                d = it.data(Qt.UserRole) or {}
                sid = str(d.get("stock_id") or "")
                if not sid or sid not in self._web_selected_ids:
                    continue
                mats.append({
                    "type": "web_stock",
                    "path": d.get("url", ""),
                    "thumb": d.get("thumb", ""),
                    "name": (d.get("title") or "").strip() or "免版权素材",
                    "provider": d.get("provider", ""),
                    "author": d.get("author", ""),
                    "stock_id": sid,
                    "media_type": (d.get("type") or "").lower(),
                    "width": d.get("width") or 0,
                    "height": d.get("height") or 0,
                    "duration_sec": d.get("duration_sec") or 0,
                })
            if mats:
                self.selected_materials = mats
                self.selected_material = mats[0]
                self.accept()


# ─────────────────────────────── Main Page ──────────────────────────────────

def _sb_server_url():
    """读取服务端统一地址（compute_server_url）。"""
    try:
        import json as _json
        import os as _os

        from config.paths import AI_CONFIG_FILE
        if _os.path.isfile(AI_CONFIG_FILE):
            cfg = _json.load(open(AI_CONFIG_FILE, encoding="utf-8"))
            url = (cfg.get("compute_server_url") or "").strip().rstrip("/")
            if url:
                return url
    except Exception:
        pass
    return ""


class _StoryboardScriptListLoader(BaseWorker):
    """从服务端 GET /api/storyboard/scripts 拉取已有分镜脚本列表摘要。"""
    finished = Signal(list)

    def do_work(self):
        from utils.storyboard_client import list_scripts
        self.finished.emit(list_scripts(page=1, page_size=100))


class _StoryboardScriptDetailLoader(BaseWorker):
    """从服务端 GET /api/storyboard/scripts/{id} 拉取完整分镜脚本。"""
    finished = Signal(dict)

    def __init__(self, script_id):
        super().__init__()
        self.script_id = script_id

    def do_work(self):
        from utils.storyboard_client import get_script
        self.finished.emit(get_script(self.script_id))


class StoryboardPage(BasePage):
    def __init__(self, parent_widget, main_window):
        super().__init__(parent_widget, main_window)
        self.worker = None
        self.shot_cards = []
        self._selected_stylization = None
        self.feishu_record = None
        self.current_product = {}   # 当前产品上下文（品牌/型号/产品类型），供素材检索使用

    # ──────────────────────────── UI ────────────────────────────────
    def setup(self):
        layout = QVBoxLayout(self.parent_widget)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(20)

        # 标题行：标题 + 介绍（介绍放标题后，stretch 推开资源监控区）
        hdr = QHBoxLayout()
        heading = QLabel(" 分镜脚本创作")
        heading.setObjectName("heading")
        hdr.addWidget(heading)
        desc = ElidedLabel("视频分镜设计 + 即梦 / MG 动画素材生成", max_lines=1)
        desc.setObjectName("muted_text")
        hdr.addWidget(desc)
        hdr.addStretch()
        layout.addLayout(hdr)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setObjectName("themeSplitter")
        left = self._build_left()
        right = self._build_right()
        # 左栏固定最小宽度，避免被压缩/遮盖（参考图像抠图页的左卡面板）
        left.setMinimumWidth(320)
        right.setMinimumWidth(420)
        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setStretchFactor(0, 0)   # 左栏不随窗口拉伸
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([360, 920])
        layout.addWidget(splitter, 1)

        status_row = QHBoxLayout()
        self.lbl_status = QLabel("")
        self.lbl_status.setObjectName("muted_text")
        status_row.addWidget(self.lbl_status)
        self.pbar = QProgressBar()
        self.pbar.setVisible(False)
        self.pbar.setRange(0, 0)
        self.pbar.setMaximumWidth(200)
        status_row.addWidget(self.pbar)
        layout.addLayout(status_row)

        self._reload_stylizations()
        self._reload_sb_scripts()

    def _build_left(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)

        container = QWidget()
        col = QVBoxLayout(container)
        col.setContentsMargins(0, 0, 10, 0)
        col.setSpacing(14)

        # ── 已有脚本：搜索选择 + 继续创作 ────────────────────────────────
        card_script = QFrame()
        card_script.setObjectName("card")
        sxp = QVBoxLayout(card_script)
        sxp.setContentsMargins(20, 16, 20, 16)
        sxp.setSpacing(10)
        sxp.addWidget(self._card_title(" 已有脚本（继续创作）"))
        scr_row = QHBoxLayout()
        self.combo_sb_script = SearchableComboBox(placeholder="搜索服务端已有脚本…")
        self.combo_sb_script.addItem("── 创建新脚本 ──", None)
        scr_row.addWidget(self.combo_sb_script, 1)
        self.btn_continue_sb = QPushButton("继续创作")
        self.btn_continue_sb.setObjectName("primary_button")
        self.btn_continue_sb.setToolTip("按选中脚本数据填充风格化、视频文案与分镜脚本")
        self.btn_continue_sb.clicked.connect(self._continue_from_script)
        scr_row.addWidget(self.btn_continue_sb)
        sxp.addLayout(scr_row)
        col.addWidget(card_script)

        # ── 风格化 ────────────────────────────────────────────
        card_style = QFrame()
        card_style.setObjectName("card")
        sp = QVBoxLayout(card_style)
        sp.setContentsMargins(20, 16, 20, 16)
        sp.setSpacing(10)

        sp.addWidget(self._card_title(" 风格化（可选）"))

        ratio_row = QHBoxLayout()
        self.combo_stylization = SearchableComboBox(placeholder="输入风格名称搜索…")
        self.combo_stylization.addItem("── 不使用风格化 ──", None)
        self.combo_stylization.currentIndexChanged.connect(self._on_stylization_selected)
        ratio_row.addWidget(self.combo_stylization, 1)
        btn_reset_style = QPushButton(" 重置")
        btn_reset_style.setObjectName("secondary_button")
        btn_reset_style.setToolTip("重新加载知识库风格化列表")
        btn_reset_style.clicked.connect(self._reload_stylizations)
        ratio_row.addWidget(btn_reset_style)
        sp.addLayout(ratio_row)

        sp.addWidget(self._muted_lbl("风格画像："))
        self.text_style_portrait = QTextEdit()
        self.text_style_portrait.setReadOnly(True)
        self.text_style_portrait.setFixedHeight(75)
        self.text_style_portrait.setPlaceholderText("选择风格化后在此显示...")
        sp.addWidget(self.text_style_portrait)

        sp.addWidget(self._muted_lbl("附加提示词（调整文案）："))
        self.edit_extra_prompt = QTextEdit()
        self.edit_extra_prompt.setFixedHeight(55)
        self.edit_extra_prompt.setPlaceholderText("可输入补充要求，配合风格化通过大模型重新调整视频文案...")
        sp.addWidget(self.edit_extra_prompt)

        self.btn_adjust_copy = QPushButton(" 大模型调整文案")
        self.btn_adjust_copy.setObjectName("secondary_button")
        self.btn_adjust_copy.clicked.connect(self._adjust_copy)
        sp.addWidget(self.btn_adjust_copy)

        col.addWidget(card_style)

        # ── 视频文案 ──────────────────────────────────────────────────
        card_copy = QFrame()
        card_copy.setObjectName("card")
        cp = QVBoxLayout(card_copy)
        cp.setContentsMargins(20, 16, 20, 16)
        cp.setSpacing(10)

        cp.addWidget(self._card_title(" 视频文案（可在此处精简或重新编辑）"))

        self.edit_copy = QTextEdit()
        self.edit_copy.setPlaceholderText(
            "请在此输入需要拆解分镜的视频文案；或由产品文案创作页点击「前往分镜设计」自动传入。"
        )
        self.edit_copy.setMinimumHeight(260)
        cp.addWidget(self.edit_copy, 1)

        self.btn_gen_sb = QPushButton("生成分镜脚本")
        self.btn_gen_sb.setObjectName("primary_button")
        self.btn_gen_sb.clicked.connect(self._generate_storyboard)
        cp.addWidget(self.btn_gen_sb)

        col.addWidget(card_copy, 1)
        scroll.setWidget(container)
        return scroll

    def _build_right(self):
        panel = QWidget()
        col = QVBoxLayout(panel)
        col.setContentsMargins(10, 0, 0, 0)
        col.setSpacing(10)

        # ── 分镜脚本头部 ──────────────────────────────────────────────
        card_sb = QFrame()
        card_sb.setObjectName("card")
        sb = QVBoxLayout(card_sb)
        sb.setContentsMargins(20, 14, 20, 14)
        sb.setSpacing(10)

        # 头信息行
        hdr = QHBoxLayout()
        hdr.addWidget(QLabel(" 分镜脚本（可直接编辑各镜头字段）"))
        hdr.addStretch()

        self.btn_auto_bind = QPushButton(" 相似度自动绑定")
        self.btn_auto_bind.setObjectName("secondary_button")
        self.btn_auto_bind.setToolTip("按每个镜头的画面描述做 CLIP 相似度检索，自动绑定最相似素材")
        self.btn_auto_bind.clicked.connect(self._auto_bind_materials)
        hdr.addWidget(self.btn_auto_bind)

        lbl_ratio = QLabel("画幅")
        lbl_ratio.setObjectName("muted_text")
        hdr.addWidget(lbl_ratio)

        self.combo_shot_ratio = QComboBox()
        self.combo_shot_ratio.addItems(["9:16", "16:9", "1:1"])
        self.combo_shot_ratio.setFixedWidth(80)
        self.combo_shot_ratio.currentTextChanged.connect(self._update_sb_header)
        hdr.addWidget(self.combo_shot_ratio)

        sb.addLayout(hdr)

        # 另起一行放画幅、时长等信息（使用橙色）
        self.lbl_sb_info = QLabel("总时长：0 s  |  0 镜  （竖屏）")
        self.lbl_sb_info.setStyleSheet("color: #ff9800; font-weight: bold; font-size: 12px; margin-left: 2px;")
        sb.addWidget(self.lbl_sb_info)

        # 镜头列表（垂直滚动，每镜头一行）
        self.sb_scroll = QScrollArea()
        self.sb_scroll.setWidgetResizable(True)
        self.sb_scroll.setFrameShape(QScrollArea.NoFrame)
        self.sb_container = QWidget()
        self.sb_col = QVBoxLayout(self.sb_container)
        self.sb_col.setSpacing(8)
        self.sb_col.setContentsMargins(0, 0, 0, 0)
        self.sb_col.addStretch()
        self.sb_scroll.setWidget(self.sb_container)
        sb.addWidget(self.sb_scroll, 1)

        # 操作行（同一行）：同步多维表格 → 创建飞书文档 → 飞书关联状态 → 保存分镜脚本
        bottom_row = QHBoxLayout()
        self.btn_sync_bitable = QPushButton(" 同步到多维表格")
        self.btn_sync_bitable.setObjectName("secondary_button")
        self.btn_sync_bitable.setEnabled(False)
        self.btn_sync_bitable.clicked.connect(lambda: self._upload_to_feishu("bitable"))
        bottom_row.addWidget(self.btn_sync_bitable)
        self.btn_sync_docx = QPushButton(" 创建飞书文档")
        self.btn_sync_docx.setObjectName("secondary_button")
        self.btn_sync_docx.setEnabled(False)
        self.btn_sync_docx.clicked.connect(lambda: self._upload_to_feishu("docx"))
        bottom_row.addWidget(self.btn_sync_docx)
        self.lbl_feishu_info = QLabel("飞书关联：无")
        self.lbl_feishu_info.setObjectName("muted_text")
        bottom_row.addWidget(self.lbl_feishu_info)
        bottom_row.addStretch()
        btn_save = QPushButton(" 保存分镜脚本")
        btn_save.setObjectName("secondary_button")
        btn_save.setToolTip("将分镜脚本（JSON + 文本）保存到素材管理目录，并同步到服务端")
        btn_save.clicked.connect(self._save_storyboard)
        bottom_row.addWidget(btn_save)
        sb.addLayout(bottom_row)

        appid, appsecret, *_ = self._get_feishu_config()
        if appid and appsecret:
            self.btn_sync_bitable.setEnabled(True)
            self.btn_sync_docx.setEnabled(True)

        col.addWidget(card_sb, 1)
        return panel

    # ──────────────────── 小部件辅助 ────────────────────────────────
    @staticmethod
    def _card_title(text):
        lbl = QLabel(text)
        lbl.setObjectName("card_title")
        return lbl

    @staticmethod
    def _muted_lbl(text):
        lbl = QLabel(text)
        lbl.setObjectName("muted_text")
        return lbl

    # ──────────────────── 数据传递 ──────────────────────────────────
    def set_copywriting(self, text, feishu_record=None, ratio=None, stylization_id=None,
                        product=None):
        self.current_product = product or {}
        self.edit_copy.setPlainText(text)
        self.feishu_record = feishu_record
        if feishu_record:
            self.lbl_feishu_info.setText(f" 已关联飞书选题: {feishu_record.get('topic', '')}")
            self.btn_sync_bitable.setEnabled(True)
            self.btn_sync_docx.setEnabled(True)
            self.lbl_status.setText("已载入飞书选题文案，请点击「生成分镜脚本」开始拆解。")
        else:
            self.lbl_feishu_info.setText("飞书关联：无")
            self.btn_sync_bitable.setEnabled(False)
            appid, appsecret, *_ = self._get_feishu_config()
            self.btn_sync_docx.setEnabled(bool(appid and appsecret))
            self.lbl_status.setText("已载入文案，请点击「生成分镜脚本」。")

        if ratio and ratio in ["9:16", "16:9", "1:1"]:
            self.combo_shot_ratio.setCurrentText(ratio)

        if stylization_id:
            for i in range(self.combo_stylization.count()):
                data = self.combo_stylization.itemData(i)
                if isinstance(data, dict) and data.get("id") == stylization_id:
                    self.combo_stylization.setCurrentIndex(i)
                    break

    def reload_sources(self):
        self._reload_stylizations()

    # ──────────────────── 保存分镜脚本 ──────────────────────────────
    def _default_storyboard_name(self):
        """从文案首行 + 日期生成默认文件名。"""
        import re as _re
        date_str = datetime.now().strftime("%Y%m%d")
        title = ""
        if self.feishu_record:
            title = self.feishu_record.get("topic", "")
        if not title:
            copy = self.edit_copy.toPlainText().strip()
            for line in copy.split("\n"):
                line = line.strip()
                if len(line) >= 4:
                    title = line[:20]
                    break
        title = _re.sub(r'[\\/:*?"<>|\r\n\t#*【】\[\]]', "_", title).strip("_") or "分镜脚本"
        return f"{date_str}_{title}"

    def _collect_shots(self):
        shots = []
        for c in self.shot_cards:
            mats = c.get("materials")
            if mats:
                mat_type = mats[0].get("type", "")
                mat_path = ",".join((m.get("path") or "") for m in mats)
                mat_hash = ",".join((m.get("hash") or "") for m in mats)
                mat_mid = next((m.get("mid") for m in mats if m.get("mid")), "")
            else:
                mat = c.get("material") or {}
                mat_type = mat.get("type", "")
                mat_path = mat.get("path", "")
                mat_hash = mat.get("hash", "")
                mat_mid = mat.get("mid", "")
            shots.append({
                "index": c["idx"],
                "shot_type": c["combo_type"].currentText(),
                "duration": c["spin_dur"].value(),
                "sfx": c["edit_sfx"].text().strip(),
                "visual": c["desc"].toPlainText().strip(),
                "narration": c["narration"].toPlainText().strip(),
                "material_type": mat_type,
                "material_path": mat_path,
                "material_hash": mat_hash,
                "material_id": mat_mid,
            })
        return shots

    def _save_storyboard(self):
        if not self.shot_cards:
            self.show_warning("当前没有分镜内容，请先生成分镜脚本。", "无内容")
            return

        default_name = self._default_storyboard_name()
        ratio = self.combo_shot_ratio.currentText()
        style_name = self._selected_stylization.get("name", "") if self._selected_stylization else ""
        shots = self._collect_shots()
        orient = {"9:16": "竖屏", "16:9": "横屏", "1:1": "方形"}.get(ratio, ratio)
        total_dur = sum(s["duration"] for s in shots)
        # 优先沿用「继续创作」加载的服务端脚本 topic（服务端按 topic 覆盖更新），
        # 其次飞书选题，最后用默认文件名
        topic = (getattr(self, "_server_script_topic", "") or ""
                 or (self.feishu_record or {}).get("topic", "") or default_name)

        # ── 保存选项对话框 ──
        dlg = QDialog(self.parent_widget)
        dlg.setWindowTitle("保存分镜脚本")
        dlg.setMinimumWidth(380)
        dl = QVBoxLayout(dlg)
        dl.setSpacing(10)

        dl.addWidget(QLabel(f"共 {len(shots)} 镜 · {total_dur}s · {orient}（{ratio}）"))

        dl.addWidget(QLabel("文件名（不含扩展名）："))
        name_edit = QLineEdit(default_name)
        dl.addWidget(name_edit)

        dl.addWidget(QLabel("导出格式："))
        bg = QButtonGroup(dlg)
        fmt_excel = QRadioButton("Excel（.xlsx）— 默认")
        fmt_md = QRadioButton("Markdown（.md）")
        fmt_json = QRadioButton("JSON（.json，供脚本成片）")
        fmt_both = QRadioButton("Excel + Markdown")
        fmt_all = QRadioButton("Excel + Markdown + JSON")
        fmt_excel.setChecked(True)
        for rb in (fmt_excel, fmt_md, fmt_json, fmt_both, fmt_all):
            bg.addButton(rb)
            dl.addWidget(rb)

        # ── 飞书同步选项 ──
        chk_feishu = QCheckBox("同步到飞书文档")
        appid, appsecret, *_ = self._get_feishu_config()
        chk_feishu.setEnabled(bool(appid and appsecret))
        chk_feishu.setToolTip("" if (appid and appsecret) else "请先在「环境配置」页配置飞书 AppID/AppSecret")
        dl.addWidget(chk_feishu)

        # ── 服务端同步选项（脚本成片直接从服务端读取）──
        server_base = _sb_server_url()
        chk_server = QCheckBox("同步到服务端（供「脚本成片」直接选择）")
        chk_server.setChecked(bool(server_base))
        chk_server.setEnabled(bool(server_base))
        chk_server.setToolTip("" if server_base else "未配置服务端地址（系统设置 → 统一计算节点地址），保存后仅留在本地")
        dl.addWidget(chk_server)

        bb = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bb.button(QDialogButtonBox.Ok).setText("保存")
        bb.button(QDialogButtonBox.Cancel).setText("取消")
        bb.accepted.connect(dlg.accept)
        bb.rejected.connect(dlg.reject)
        dl.addWidget(bb)

        if dlg.exec() != QDialog.Accepted:
            return

        base_name = name_edit.text().strip() or default_name
        do_excel = fmt_excel.isChecked() or fmt_both.isChecked() or fmt_all.isChecked()
        do_md = fmt_md.isChecked() or fmt_both.isChecked() or fmt_all.isChecked()
        do_json = fmt_json.isChecked() or fmt_all.isChecked()
        do_feishu = chk_feishu.isChecked()
        do_server = chk_server.isChecked()

        # 保存到配置的素材目录（与浏览器下载目录对齐）
        import re as _re
        safe_topic = _re.sub(r'[\\/:*?"<>|\r\n\t]', "_", topic)[:40]
        out_dir = os.path.join(KNOWLEDGE_MEDIA_DIR, safe_topic, "storyboard")
        os.makedirs(out_dir, exist_ok=True)

        saved_files = []

        if do_excel:
            xlsx_path = os.path.join(out_dir, base_name + ".xlsx")
            self._export_storyboard_excel(xlsx_path, topic, ratio, orient, style_name, total_dur, shots)
            saved_files.append(xlsx_path)

        if do_md:
            md_path = os.path.join(out_dir, base_name + ".md")
            self._export_storyboard_md(md_path, topic, ratio, orient, style_name, total_dur, shots)
            saved_files.append(md_path)

        if do_json:
            json_path = os.path.join(out_dir, base_name + ".json")
            self._export_storyboard_json(json_path, topic, ratio, total_dur, shots)
            saved_files.append(json_path)

        if not saved_files:  # fallback
            xlsx_path = os.path.join(out_dir, base_name + ".xlsx")
            self._export_storyboard_excel(xlsx_path, topic, ratio, orient, style_name, total_dur, shots)
            saved_files.append(xlsx_path)

        # 供「一键成片 → 脚本成片」使用：无论选择哪种格式都额外生成 JSON
        # （一键成片只扫描 KNOWLEDGE_MEDIA_DIR/<选题>/storyboard/*.json）
        json_path = os.path.join(out_dir, base_name + ".json")
        if not os.path.exists(json_path):
            self._export_storyboard_json(json_path, topic, ratio, total_dur, shots)
        if json_path not in saved_files:
            saved_files.append(json_path)

        if do_feishu:
            self._upload_to_feishu("docx")

        # 同步到服务端（异步 POST /api/storyboard/scripts，同 topic 覆盖更新）
        server_note = ""
        if do_server:
            self._upload_storyboard_to_server(topic, ratio, total_dur, shots)
            server_note = "\n\n（已发起同步到服务端，状态见页面底部提示；\n可在「一键成片 → 脚本成片」刷新后直接选择）"


        self.show_info(
            f"分镜脚本已保存至：\n{out_dir}\n\n"
            + "\n".join(os.path.basename(f) for f in saved_files)
            + server_note,
            "保存成功",
        )

    def _upload_storyboard_to_server(self, topic, ratio, total_dur, shots):
        """把分镜脚本上传到服务端（POST /api/storyboard/scripts，异步）。

        服务端契约（http://<server>:8000/openapi.json 的 ScriptIn/Shot）：
          保存字段 topic/ratio/total_duration/shot_count/shots/saved_at/product/
          brand/model/category/name；shots 元素 {index, shot_type, visual, audio,
          sfx, duration, material_path, material_type, material_hash, material_id}；
          客户端本地字段 narration → 服务端 audio。
        同 topic 重复保存视为更新（不新增）。
        返回 True 表示已发起且请求成功；失败仅记日志，不影响本地保存。
        """
        from utils.thread_worker import TaskWorker as Worker
        base = _sb_server_url()
        if not base:
            log.warning("[分镜脚本] 未配置服务端地址，跳过上传")
            return False

        # 字段对齐：narration → audio（服务端 storyboard_montage 契约）
        server_shots = []
        for sh in shots or []:
            mid = sh.get("material_id") or ""
            try:
                mid = int(mid)
            except (ValueError, TypeError):
                mid = 0
            server_shots.append({
                "index": sh.get("index", 0),
                "shot_type": sh.get("shot_type", ""),
                "duration": sh.get("duration", 3),
                "visual": sh.get("visual", ""),
                "audio": sh.get("audio", "") or sh.get("narration", ""),
                "sfx": sh.get("sfx", ""),
                "material_path": sh.get("material_path", ""),
                "material_type": sh.get("material_type", ""),
                "material_hash": sh.get("material_hash", ""),
                "material_id": mid,
            })
        prod = getattr(self, "current_product", {}) or {}
        payload = {
            "topic": topic or "未命名分镜脚本",
            "ratio": ratio,
            "total_duration": float(total_dur or 0),
            "shot_count": len(server_shots),
            "shots": server_shots,
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            # 产品上下文（ScriptIn.product + 顶层品牌字段），供素材检索/一键成片使用
            "product": {
                "brand": str(prod.get("brand") or ""),
                "model": str(prod.get("model") or ""),
                "category": str(prod.get("category") or ""),
                "name": str(prod.get("name") or ""),
            },
            "brand": str(prod.get("brand") or ""),
            "model": str(prod.get("model") or ""),
            "category": str(prod.get("category") or ""),
            "name": str(prod.get("name") or ""),
        }

        def _do():
            from utils.storyboard_client import save_script
            ok = save_script(payload)
            if ok:
                log.info(f"[分镜脚本] 已上传服务端 topic={payload['topic']} shots={len(server_shots)}")
            else:
                log.warning(f"[分镜脚本] 上传失败 topic={payload['topic']}")
            return ok

        def _done(ok):
            if not ok:
                self.lbl_status.setText("注意： 分镜脚本已保存到本地，但同步服务端失败")
                log.warning("[分镜脚本] 同步服务端失败（详见日志）")
            else:
                self.lbl_status.setText("完成： 分镜脚本已保存并同步到服务端")
                # 刷新「继续创作」下拉，服务端脚本立即可见
                try:
                    self._reload_sb_scripts()
                except Exception:
                    pass
                # 同时刷新「一键成片 → 脚本成片」的脚本列表
                try:
                    compile_tool = getattr(self.main_window, "compile_video_tool", None)
                    if compile_tool and hasattr(compile_tool, "_populate_scripts"):
                        compile_tool._populate_scripts()
                except Exception:
                    pass

        def _err(e):
            self.lbl_status.setText("注意： 分镜脚本已保存到本地，但同步服务端失败")
            log.warning(f"[分镜脚本] 同步服务端异常: {e}")

        w = Worker(_do)
        w.finished.connect(_done)
        w.error.connect(_err)
        self.track_worker(w)
        w.start()
        return True

    def _export_storyboard_excel(self, path, topic, ratio, orient, style_name, total_dur, shots):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            raise RuntimeError("缺少 openpyxl，请运行：pip install openpyxl")

        wb = openpyxl.Workbook()

        # Sheet 1：概览
        ws_info = wb.active
        ws_info.title = "概览"
        header_font = Font(bold=True, size=12)
        ws_info.column_dimensions["A"].width = 16
        ws_info.column_dimensions["B"].width = 40
        rows_info = [
            ("选题", topic), ("画幅", f"{orient}（{ratio}）"),
            ("风格化", style_name or "—"), ("总时长", f"{total_dur}s"),
            ("镜头数", len(shots)), ("生成日期", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        for r, (k, v) in enumerate(rows_info, 1):
            ws_info.cell(r, 1, k).font = header_font
            ws_info.cell(r, 2, str(v))

        # Sheet 2：分镜脚本
        ws = wb.create_sheet("分镜脚本")
        cols = ["镜号", "镜别", "时长(s)", "音效", "画面描述", "旁白/台词", "引用素材", "素材Hash"]
        widths = [6, 8, 8, 16, 40, 40, 30, 20]
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = Font(bold=True, color="FFFFFF")
        for ci, (col, w) in enumerate(zip(cols, widths, strict=True), 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
            cell = ws.cell(1, ci, col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        alt_fill = PatternFill("solid", fgColor="F5F7FA")
        for ri, s in enumerate(shots, 2):
            row_data = [
                s["index"], s["shot_type"], s["duration"],
                s["sfx"], s["visual"], s["narration"], s["material_path"], s.get("material_hash", ""),
            ]
            fill = alt_fill if ri % 2 == 0 else None
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(ri, ci, val)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    cell.fill = fill
            ws.row_dimensions[ri].height = max(40, len(str(s["visual"])) // 3)

        wb.save(path)

    def _export_storyboard_md(self, path, topic, ratio, orient, style_name, total_dur, shots):
        lines = [
            f"# 分镜脚本 — {topic}",
            "",
            f"**画幅**：{orient}（{ratio}）　**风格**：{style_name or '—'}　**总时长**：{total_dur}s　**镜头数**：{len(shots)}",
            "",
            f"*生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}*",
            "",
            "---",
            "",
            "| 镜号 | 镜别 | 时长 | 音效 | 画面描述 | 旁白/台词 | 引用素材 | 素材Hash |",
            "|:---:|:---:|:---:|------|---------|-----------|---------|---------|",
        ]
        for s in shots:
            def _esc(v):
                return str(v).replace("|", "｜").replace("\n", " ").strip()
            lines.append(
                f"| {s['index']} | {_esc(s['shot_type'])} | {s['duration']}s "
                f"| {_esc(s['sfx']) or '—'} | {_esc(s['visual'])} "
                f"| {_esc(s['narration']) or '—'} | {_esc(s['material_path']) or '—'} | {_esc(s.get('material_hash','')) or '—'} |"
            )
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    def _export_storyboard_json(self, path, topic, ratio, total_dur, shots):
        """导出为 JSON（供「一键成片 > 脚本成片」tab 选择并提交服务端）。
        结构：{topic, ratio, total_duration, shot_count, shots:[{index,shot_type,duration,sfx,visual,narration,material_type,material_path,material_hash}], saved_at}
        """
        import json as _json
        import time as _time
        data = {
            "topic": topic,
            "ratio": ratio,
            "total_duration": total_dur,
            "shot_count": len(shots),
            "shots": shots,   # _collect_shots() 的输出，已含素材路径+文案
            "saved_at": int(_time.time()),
        }
        with open(path, "w", encoding="utf-8") as f:
            _json.dump(data, f, ensure_ascii=False, indent=2)

    # ──────────────────── 风格化 ────────────────────────────────────
    def _reload_sb_scripts(self):
        """从服务端加载已有分镜脚本，填充「继续创作」下拉。"""
        w = self.track_worker(_StoryboardScriptListLoader())
        w.finished.connect(self._on_sb_scripts_loaded)
        w.error.connect(self._on_sb_scripts_load_error)
        w.start()

    def _on_sb_scripts_load_error(self, msg):
        log.warning(f"加载已有脚本失败: {msg}")
        hint = "服务端不可达或未实现接口"
        if "404" in str(msg):
            hint = "服务端未实现分镜脚本接口（404）"
        elif "无法连接" in str(msg):
            hint = "服务端不可达，请检查统一计算节点地址"
        self.lbl_status.setText(f"注意： 加载服务端脚本失败：{hint}（{msg}）")

    def _on_sb_scripts_loaded(self, items):
        cur = self.combo_sb_script.currentData()
        self.combo_sb_script.blockSignals(True)
        self.combo_sb_script.clear()
        self.combo_sb_script.addItem("── 创建新脚本 ──", None)
        for it in items or []:
            sid = it.get("id")
            if not sid:
                continue
            label = f"[{it.get('topic', '')}] {it.get('shot_count', 0)}镜"
            if it.get("saved_at"):
                label += f" · {it['saved_at']}"
            self.combo_sb_script.addItem(label, {"id": sid, "topic": it.get("topic", "")})
        self.combo_sb_script.blockSignals(False)
        if cur and isinstance(cur, dict):
            for i in range(self.combo_sb_script.count()):
                d = self.combo_sb_script.itemData(i)
                if isinstance(d, dict) and d.get("id") == cur.get("id"):
                    self.combo_sb_script.setCurrentIndex(i)
                    break

    def _continue_from_script(self):
        sel = self.combo_sb_script.currentData()
        if not sel or not sel.get("id"):
            # 创建新脚本：清空当前内容
            self.edit_copy.clear()
            self._render_shots([])
            self._server_script_topic = ""
            self.lbl_status.setText("已切换到「创建新脚本」模式。")
            return
        w = self.track_worker(_StoryboardScriptDetailLoader(sel["id"]))
        w.finished.connect(self._apply_server_script)
        w.error.connect(lambda e: self.show_error(f"加载脚本失败：{e}"))
        w.start()
        self.lbl_status.setText("正在加载脚本…")

    def _apply_server_script(self, script):
        if not script:
            return
        shots = script.get("shots") or []
        # 视频文案：镜头旁白拼接
        copy_lines = [str(sh.get("audio") or "").strip() for sh in shots if (sh.get("audio") or "").strip()]
        self.edit_copy.setPlainText(chr(10).join(copy_lines) if copy_lines else "")
        # 画幅
        ratio = script.get("ratio") or ""
        if ratio in ("9:16", "16:9", "1:1"):
            self.combo_shot_ratio.setCurrentText(ratio)
        # 分镜脚本
        self._render_shots(shots)
        # 恢复产品上下文（ScriptIn.product 优先，顶层品牌字段兜底）——
        # 引用素材检索 / 相似度自动绑定会自动带上品牌/型号/品类
        prod = script.get("product") or {}
        if not isinstance(prod, dict):
            prod = {}
        self.current_product = {
            "brand": str(prod.get("brand") or script.get("brand") or ""),
            "model": str(prod.get("model") or script.get("model") or ""),
            "category": str(prod.get("category") or script.get("category") or ""),
            "name": str(prod.get("name") or script.get("name") or ""),
        }
        self._server_script_topic = (script.get("topic") or "").strip()
        self.lbl_status.setText(f"已载入脚本「{script.get('topic', '')}」（{len(shots)} 镜），可继续编辑并生成。")

    def _reload_stylizations(self):
        mgr = MyKnowledgeManager()
        items = [it for it in mgr.items if it.get("type") == STYLIZATION_TYPE]
        items.sort(key=lambda x: x.get("score", 0), reverse=True)
        self.combo_stylization.blockSignals(True)
        prev_id = (self._selected_stylization or {}).get("id")
        self.combo_stylization.clear()
        self.combo_stylization.addItem("── 不使用风格化 ──", None)
        restore_idx = 0
        for it in items:
            label = it.get("name", "（无名）")
            score = it.get("score", 0)
            if score:
                label += f"  {score}"
            self.combo_stylization.addItem(label, it)
            if it.get("id") == prev_id:
                restore_idx = self.combo_stylization.count() - 1
        self.combo_stylization.setCurrentIndex(restore_idx)
        self.combo_stylization.blockSignals(False)
        self._on_stylization_selected()

    def _on_stylization_selected(self):
        data = self.combo_stylization.currentData()
        self._selected_stylization = data if isinstance(data, dict) else None
        self.text_style_portrait.setPlainText(
            self._selected_stylization.get("content", "") if self._selected_stylization else ""
        )

    # ──────────────────── 文案调整 ──────────────────────────────────
    def _adjust_copy(self):
        copy_text = self.edit_copy.toPlainText().strip()
        if not copy_text:
            QMessageBox.warning(self.parent_widget, "文案为空", "请先填写视频文案。")
            return
        extra = self.edit_extra_prompt.toPlainText().strip()
        style_text = (self._selected_stylization.get("content") or "").strip() if self._selected_stylization else ""
        parts = [f"原始视频文案：\n{copy_text}"]
        if style_text:
            parts.append(f"风格化要求（HOW to write）：\n{style_text}")
        if extra:
            parts.append(f"额外要求：\n{extra}")
        parts.append("请根据以上要求重新输出优化后的视频文案，保持核心信息不变，只调整表达风格和措辞。")
        self._run_llm(
            "你是专业的短视频文案创作者，根据用户要求对视频文案进行优化和调整。",
            "\n\n".join(parts),
            lambda c: (self.edit_copy.setPlainText(c),
                       self.lbl_status.setText("文案调整完成，可继续生成分镜脚本。")),
            self.btn_adjust_copy, "AI 正在调整文案…",
        )

    # ──────────────────── 分镜卡片 ──────────────────────────────────
    def _make_shot_card(self, idx, shot_type="近景", visual="", audio="", sfx="", duration=5):
        frame = QFrame()
        frame.setObjectName("card")
        v = QVBoxLayout(frame)
        v.setContentsMargins(14, 10, 14, 10)
        v.setSpacing(6)

        # 头部行：镜头号 镜别 时长 音效 [引用素材]
        hdr = QHBoxLayout()
        hdr.addWidget(QLabel(f"<b>镜头 {idx}</b>"))

        hdr.addWidget(QLabel("镜别"))
        combo_type = QComboBox()
        combo_type.addItems(SHOT_TYPES)
        if shot_type in SHOT_TYPES:
            combo_type.setCurrentText(shot_type)
        combo_type.setFixedWidth(70)
        hdr.addWidget(combo_type)

        hdr.addSpacing(8)
        hdr.addWidget(QLabel("时长"))
        spin_dur = QSpinBox()
        spin_dur.setRange(1, 120)
        spin_dur.setValue(int(duration) if duration else 5)
        spin_dur.setSuffix(" s")
        spin_dur.setFixedWidth(65)
        spin_dur.valueChanged.connect(self._update_sb_header)
        hdr.addWidget(spin_dur)

        hdr.addSpacing(8)
        hdr.addWidget(QLabel("音效"))
        edit_sfx = QLineEdit()
        edit_sfx.setText(sfx)
        edit_sfx.setPlaceholderText("如：轻松背景音乐、键盘敲击声…")
        hdr.addWidget(edit_sfx, 1)  # 音效输入框占据剩余空间，按钮紧随其后贴在右侧
        btn_mat = QPushButton(" 引用素材")
        btn_mat.setObjectName("secondary_button")
        # 不设固定高度：跟随全局 QSS（与底部「保存分镜脚本」等按钮等高）
        hdr.addWidget(btn_mat)
        v.addLayout(hdr)

        # 素材行：显示已绑定素材信息
        mat_row = QHBoxLayout()
        mat_lbl = QLabel("")
        mat_lbl.setObjectName("muted_text")
        mat_row.addWidget(mat_lbl, 1)
        v.addLayout(mat_row)

        # 内容行：画面描述 + 旁白（左） | 缩略图（右）
        content_row = QHBoxLayout()
        content_row.setSpacing(10)

        text_col = QVBoxLayout()
        text_col.setSpacing(4)
        text_col.addWidget(self._muted_lbl("画面描述"))
        desc = QTextEdit()
        desc.setPlainText(visual)
        desc.setFixedHeight(58)
        desc.setPlaceholderText("镜头画面内容描述（即梦出图提示词，可直接编辑）")
        text_col.addWidget(desc)
        text_col.addWidget(self._muted_lbl("旁白台词"))
        narration = QTextEdit()
        narration.setPlainText(audio)
        narration.setFixedHeight(38)
        narration.setPlaceholderText("旁白/台词")
        text_col.addWidget(narration)
        content_row.addLayout(text_col, 1)

        thumb = QLabel("未出图")
        thumb.setObjectName("muted_text")
        # 只锁宽度，高度随行拉伸与左侧文本列对齐，避免预览区被截断
        thumb.setFixedWidth(120)
        thumb.setMinimumHeight(130)
        thumb.setAlignment(Qt.AlignCenter)
        thumb.setStyleSheet("border:1px solid #3a3a3a; border-radius:4px;")
        content_row.addWidget(thumb, 0)

        v.addLayout(content_row)

        card = {
            "frame": frame, "idx": idx,
            "combo_type": combo_type, "spin_dur": spin_dur,
            "edit_sfx": edit_sfx, "desc": desc, "narration": narration,
            "thumb": thumb, "file": "", "material": None, "mat_lbl": mat_lbl,
        }
        btn_mat.clicked.connect(lambda: self._open_material_dialog(card))
        return card

    def _open_material_dialog(self, card):
        try:
            log.info(f"[引用素材] 按钮被点击，镜头 idx={card.get('idx')}")
            shot_desc = card["desc"].toPlainText().strip()
            ratio = self.combo_shot_ratio.currentText()
            prod = getattr(self, "current_product", {}) or {}
            # 产品上下文兜底：飞书选题 + 视频文案首句，保证检索带上产品信息
            extra = []
            topic = (getattr(self, "feishu_record", {}) or {}).get("topic", "")
            if topic:
                extra.append(str(topic)[:60])
            copy = self.edit_copy.toPlainText().strip()
            if copy:
                extra.append(copy[:120])
            log.info(f"[引用素材] 创建对话框: desc_len={len(shot_desc)} brand={prod.get('brand')} model={prod.get('model')}")
            dlg = ShotMaterialDialog(
                shot_desc=shot_desc, ratio=ratio,
                brand=str(prod.get("brand") or ""),
                model=str(prod.get("model") or ""),
                category=str(prod.get("category") or ""),
                shot_type=card["combo_type"].currentText(),
                extra_ctx="，".join(extra),
                style=(self.combo_stylization.currentText() if self._selected_stylization else ""),
                topic=str(topic),
                main_window=self.main_window,
                parent=self.parent_widget,
            )
            log.info(f"[引用素材] 对话框已创建，准备 exec()")
            result = dlg.exec()
            log.info(f"[引用素材] 对话框关闭，result={result}")
            if result == QDialog.Accepted:
                mats = getattr(dlg, "selected_materials", None) or (
                    [dlg.selected_material] if dlg.selected_material else [])
                if mats:
                    log.info(f"[引用素材] 选中 {len(mats)} 个素材")
                    self._bind_materials(card, mats)
        except Exception as e:
            log.exception(f"[引用素材] 对话框打开失败: {e}")
            self.show_error(f"引用素材对话框打开失败：{e}", "错误")

    def _set_shot_thumb(self, card, mat):
        """把素材缩略图显示到镜头卡片的右侧区域（本地文件直读，服务端素材异步下载）。"""
        thumb = card.get("thumb")
        if thumb is None or not mat:
            return
        path = (mat.get("path") or "").strip()
        tw, th = max(thumb.width(), 120), max(thumb.height(), 140)
        # 1) 本地磁盘文件（即梦生成图 / 本地素材）：直接加载
        if os.path.isfile(path):
            pm = QPixmap(path)
            if not pm.isNull():
                thumb.setPixmap(pm.scaled(tw, th, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            return
        # 2) 服务端素材（/material/serve?material_id=…）：异步拉取缩略图
        mid = str(mat.get("material_id") or mat.get("mid") or "")
        if not mid and "material_id=" in path:
            mid = path.split("material_id=", 1)[1].split("&", 1)[0]
        if not mid:
            return

        def on_thumb(_mid, data):
            pm = QPixmap()
            if data and pm.loadFromData(data) and not pm.isNull():
                thumb.setPixmap(pm.scaled(tw, th, Qt.KeepAspectRatio, Qt.SmoothTransformation))

        w = _ThumbWorker(mid)
        w.finished.connect(on_thumb)
        w.error.connect(lambda e: log.debug(f"[分镜] 镜头缩略图加载失败: {e}"))
        self.track_worker(w)
        w.start()

    def _bind_materials(self, card, mats):
        """把勾选的素材绑定到镜头：Hash 全部记录，卡片显示名称/数量。"""
        card["materials"] = list(mats)
        card["material"] = mats[0]
        for mat in mats:
            if mat.get("type") == "dreamina":
                card["file"] = mat.get("path", "")
                break
        self._set_shot_thumb(card, mats[0])
        names = "、".join((m.get("name") or "")[:10] for m in mats if m.get("name"))
        if len(mats) > 1:
            label = f"{names[:22]}{'…' if len(names) > 22 else ''}（{len(mats)}个）"
        else:
            label = names or "已绑定素材"
        card["mat_lbl"].setText(label)
        card["mat_lbl"].setToolTip("\n".join(m.get("path", "") for m in mats))

    # ── 相似度自动绑定（全部镜头）────────────────────────────────────
    def _auto_bind_materials(self):
        if not self.shot_cards:
            self.show_warning("当前没有分镜内容，请先生成分镜脚本。", "无内容")
            return

        # 收集每个镜头的检索文案：景别 + 产品上下文 + 画面描述/旁白
        prod = getattr(self, "current_product", {}) or {}
        prod_ctx = " ".join(x for x in (
            str(prod.get("brand") or ""), str(prod.get("model") or ""),
            str(prod.get("category") or "")) if x)
        shots = []
        for c in self.shot_cards:
            desc = c["desc"].toPlainText().strip()
            narr = c["narration"].toPlainText().strip()
            shot_type = c["combo_type"].currentText().strip()
            query = " ".join(x for x in (shot_type, prod_ctx, (desc or narr)) if x)
            if query:
                shots.append((c["idx"], query))

        if not shots:
            self.show_warning("所有镜头都没有画面描述/旁白，无法做相似度检索。", "无可用文案")
            return

        self.btn_auto_bind.setEnabled(False)
        self.pbar.setVisible(True)
        self.pbar.setRange(0, len(shots))
        self.pbar.setValue(0)
        self.lbl_status.setText("正在按镜头文案做相似度自动绑定…")

        self._bind_worker = _AutoBindShotsWorker(
            shots, min_score=0.0,
            filter_brand=str(prod.get("brand") or ""),
            filter_category=str(prod.get("category") or ""))
        self._bind_worker.progress.connect(self._on_auto_bind_progress)
        self._bind_worker.finished.connect(self._on_auto_bind_finished)
        self._bind_worker.error.connect(self._on_auto_bind_error)
        self._bind_worker.start()

    def _on_auto_bind_progress(self, done, total):
        self.pbar.setRange(0, total)
        self.pbar.setValue(done)
        self.lbl_status.setText(f"相似度自动绑定中… {done}/{total}")

    def _on_auto_bind_error(self, msg):
        self.btn_auto_bind.setEnabled(True)
        self.pbar.setVisible(False)
        self.lbl_status.setText(f"自动绑定失败：{msg}")

    def _on_auto_bind_finished(self, result):
        self.btn_auto_bind.setEnabled(True)
        self.pbar.setVisible(False)
        bound = 0
        for c in self.shot_cards:
            mat = result.get(c["idx"])
            if not mat:
                continue
            c["material"] = mat
            name = mat.get("name", "") or ""
            score = float(mat.get("score", 0) or 0)
            c["mat_lbl"].setText((name[:14] + ("…" if len(name) > 14 else "")) + f" {score*100:.0f}%")
            self._set_shot_thumb(c, mat)
            c["mat_lbl"].setToolTip(
                f"{mat.get('path','')}\n相似度: {score*100:.1f}%\nHash: {mat.get('hash','')}"
            )
            bound += 1
        total = len(self.shot_cards)
        self.lbl_status.setText(f"相似度自动绑定完成：{bound}/{total} 个镜头已绑定素材。")
        self.show_message(
            f"相似度自动绑定完成！\n已绑定 {bound} / {total} 个镜头。\n"
            f"未绑定的镜头多为无相似素材，可手动点「引用素材」补充。"
        )

    def _render_shots(self, shots):
        for c in self.shot_cards:
            c["frame"].setParent(None)
        self.shot_cards = []
        # remove stretch
        item = self.sb_col.takeAt(self.sb_col.count() - 1)
        del item
        for i, shot in enumerate(shots):
            try:
                idx = int(shot.get("index", i + 1))
            except (ValueError, TypeError):
                idx = i + 1
            card = self._make_shot_card(
                idx,
                shot_type=str(shot.get("shot_type", "近景")),
                visual=str(shot.get("visual", "")),
                audio=str(shot.get("audio", "")),
                sfx=str(shot.get("sfx", "")),
                duration=shot.get("duration", 5),
            )
            # 恢复素材绑定（服务端 Shot 支持 material_path/material_id）：
            # 只恢复第一个素材，缩略图与「引用素材」标签即可还原
            mp = str(shot.get("material_path") or "").strip()
            if mp:
                self._bind_materials(card, [{
                    "type": str(shot.get("material_type") or "local"),
                    "path": mp.split(",", 1)[0].strip(),
                    "name": "",
                    "hash": str(shot.get("material_hash") or "").split(",", 1)[0],
                    "mid": str(shot.get("material_id") or ""),
                }])
            self.sb_col.addWidget(card["frame"])
            self.shot_cards.append(card)
        self.sb_col.addStretch()
        self._update_sb_header()

    def _update_sb_header(self):
        total = sum(c["spin_dur"].value() for c in self.shot_cards)
        ratio = self.combo_shot_ratio.currentText()
        orient = {"9:16": "竖屏", "16:9": "横屏", "1:1": "方形"}.get(ratio, ratio)
        self.lbl_sb_info.setText(f"总时长：{total} s  |  {len(self.shot_cards)} 镜  （{orient}）")

    def _on_shot_done(self, idx, file_path):
        for c in self.shot_cards:
            if c["idx"] == idx:
                pm = QPixmap(file_path)
                if not pm.isNull():
                    c["thumb"].setPixmap(pm.scaled(120, 90, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                c["file"] = file_path
                break

    # ──────────────────── LLM 调用 ──────────────────────────────────
    def _ai_cfg(self):
        ai = getattr(self.main_window, "ai_config", {}) or {}
        model = ai.get("llm_model", "deepseek-v4-flash")
        if not model:
            QMessageBox.warning(self.parent_widget, "大模型未配置",
                                "请先在「AI 设置 / 大模型配置」中选择模型名称。")
            return None
        return "", "", model

    def _run_llm(self, system_prompt, user_prompt, on_done, busy_btn, busy_text):
        cfg = self._ai_cfg()
        if not cfg:
            return
        url, key, model = cfg
        busy_btn.setEnabled(False)
        self.lbl_status.setText(busy_text)
        self.pbar.setVisible(True)
        self.worker = LLMWorker(url, key, model, system_prompt, user_prompt)

        def _done(content):
            busy_btn.setEnabled(True)
            self.pbar.setVisible(False)
            on_done(content.strip())

        def _err(msg):
            busy_btn.setEnabled(True)
            self.pbar.setVisible(False)
            self.lbl_status.setText("生成失败。")
            QMessageBox.critical(self.parent_widget, "大模型异常", f"请求失败：\n{msg}")

        self.worker.finished.connect(_done)
        self.worker.error.connect(_err)
        self.worker.start()

    # ──────────────────── 分镜生成 ──────────────────────────────────
    def _generate_storyboard(self):
        copy_text = self.edit_copy.toPlainText().strip()
        if not copy_text:
            QMessageBox.warning(self.parent_widget, "文案为空", "请先填写或生成视频文案。")
            return

        ratio = self.combo_shot_ratio.currentText()
        orient = {"9:16": "竖屏（9:16）", "16:9": "横屏（16:9）", "1:1": "方形（1:1）"}.get(ratio, ratio)

        system_prompt = (
            "你是专业短视频导演，把视频文案拆解为专业分镜脚本。"
            "每个镜头需要包含镜别、画面描述、旁白台词、音效建议、建议时长（秒）。"
        )
        user_prompt = (
            f"请把以下短视频文案拆解为分镜脚本（{orient}画幅），约 9 个镜头，"
            "以 JSON 数组输出，每个元素含以下字段：\n"
            '  "index"(整型镜头序号), "shot_type"(镜别，如特写/近景/中景/远景/全景/俯拍/仰拍/主观/空镜), '
            '"visual"(画面描述，可作为即梦出图提示词), "audio"(旁白/台词), '
            '"sfx"(音效建议，如无则留空字符串), "duration"(建议时长秒数，整型)。\n'
            "严格只输出 JSON 数组，不要 ```json 包裹。\n\n文案：\n" + copy_text
        )
        self._run_llm(system_prompt, user_prompt, self._fill_storyboard,
                      self.btn_gen_sb, "AI 正在拆解分镜…")

    def _fill_storyboard(self, content):
        text = content.strip()
        if text.startswith("```"):
            text = text.strip("`")
            if text.lower().startswith("json"):
                text = text[4:]
            text = text.strip()
        try:
            shots = json.loads(text)
            if not isinstance(shots, list):
                raise ValueError("非 JSON 数组")
        except Exception as e:
            log.error(f"分镜 JSON 解析失败: {e}")
            self.lbl_status.setText("分镜解析失败，已将原始结果放入第一格。")
            self._render_shots([{"index": 1, "visual": content, "audio": "", "sfx": "", "duration": 5}])
            return
        self._render_shots(shots)
        total = sum(s.get("duration", 5) for s in shots if isinstance(s.get("duration"), (int, float)))
        self.lbl_status.setText(
            f"分镜已生成（{len(shots)} 个镜头，约 {total}s），可直接编辑各镜头字段，或点击「引用素材」关联本地/即梦/联网素材。"
        )

    def _open_mg(self):
        try:
            # MG动画在素材生成页（index 31）的内部 Tab2
            self.main_window.switch_page(31)
            if hasattr(self.main_window, "switch_dreamina_tab"):
                self.main_window.switch_dreamina_tab(2)
            tool = getattr(self.main_window, "mg_animation_tool", None)
            if tool and self.shot_cards and hasattr(tool, "set_default_text"):
                tool.set_default_text(self.shot_cards[0]["desc"].toPlainText().strip()[:20])
        except Exception as e:
            self.show_error(f"跳转失败：{e}")

    # ──────────────────── 即梦批量生成 ─────────────────────────────
    def _generate_shot_images(self):
        shots = [(c["idx"], c["desc"].toPlainText().strip()) for c in self.shot_cards
                 if c["desc"].toPlainText().strip()]
        if not shots:
            self.show_warning("请先生成分镜脚本。")
            return
        ratio = self.combo_shot_ratio.currentText()
        out_dir = os.path.join(DREAMINA_OUTPUT_DIR, "shots_" + datetime.now().strftime("%Y%m%d_%H%M%S"))
        self.btn_gen_shots.setEnabled(False)
        self.pbar.setVisible(True)
        self.lbl_status.setText(f"开始批量生成 {len(shots)} 个镜头素材（{ratio}）…")
        worker = BatchShotImageWorker(shots, ratio, out_dir)
        worker.phase.connect(self.lbl_status.setText)
        worker.shot_done.connect(self._on_shot_done)

        def done(files):
            self.btn_gen_shots.setEnabled(True)
            self.pbar.setVisible(False)
            if files:
                self.lbl_status.setText(f" 已生成 {len(files)} 张镜头素材：{out_dir}")
            else:
                self.lbl_status.setText("未生成任何素材（可能未登录或全部失败）。")

        worker.finished.connect(done)
        worker.error.connect(lambda e: (
            self.btn_gen_shots.setEnabled(True),
            self.pbar.setVisible(False),
            self.show_error(str(e), "镜头素材生成失败"),
        ))
        self.track_worker(worker)
        worker.start()

    # ──────────────────── 飞书同步 ──────────────────────────────────
    def _get_feishu_config(self):
        """从 CONFIG_INI_FILE 读取飞书配置。"""
        import configparser
        config = configparser.ConfigParser()
        appid = appsecret = apptoken = tableid = ""
        topicfield = "选题"
        scriptfield = "脚本"
        foldertoken = ""
        try:
            config.read(CONFIG_INI_FILE, encoding="utf-8")
            if config.has_section("Feishu"):
                appid = config.get("Feishu", "AppId", fallback="")
                appsecret = config.get("Feishu", "AppSecret", fallback="")
                apptoken = config.get("Feishu", "AppToken", fallback="")
                tableid = config.get("Feishu", "TableId", fallback="")
                topicfield = config.get("Feishu", "TopicField", fallback="选题")
                scriptfield = config.get("Feishu", "ScriptField", fallback="脚本")
                foldertoken = config.get("Feishu", "FolderToken", fallback="")
        except Exception:
            pass
        return appid, appsecret, apptoken, tableid, topicfield, scriptfield, foldertoken

    def _get_script_table_as_text(self):
        """将当前分镜脚本格式化为文本，供飞书同步。"""
        lines = []
        for card in self.shot_cards:
            idx = card["idx"]
            shot_type = card["combo_type"].currentText()
            duration = card["spin_dur"].value()
            sfx = card["edit_sfx"].text().strip()
            visual = card["desc"].toPlainText().strip()
            narration = card["narration"].toPlainText().strip()
            mat = card.get("material") or {}
            mat_path = mat.get("path", "")
            lines.append(
                f"【镜头{idx}】{shot_type} | {duration}s | 音效：{sfx or '—'}\n"
                f"画面：{visual}\n"
                f"旁白：{narration}\n"
                f"素材：{mat_path}\n"
            )
        return "\n".join(lines)

    def _upload_to_feishu(self, mode):
        """将分镜脚本同步到飞书（bitable 或 docx）。"""
        appid, appsecret, apptoken, tableid, topicfield, scriptfield, foldertoken = self._get_feishu_config()
        if not appid or not appsecret:
            QMessageBox.warning(self.parent_widget, "配置未完成",
                                "请先在「环境配置」页配置飞书 AppID 和 AppSecret。")
            return
        record_id = None
        topic_name = "新分镜脚本"
        if self.feishu_record:
            record_id = self.feishu_record.get("id")
            topic_name = self.feishu_record.get("topic", "新分镜脚本")
        if mode == "bitable" and (not record_id or not apptoken or not tableid):
            QMessageBox.warning(self.parent_widget, "无法同步",
                                "飞书多维表格配置不完整或未关联选题。\n请先在 AI 文案页同步选题。")
            return
        script_text = self._get_script_table_as_text()
        self.lbl_status.setText(f"正在同步到飞书（{mode}）…")
        self.pbar.setVisible(True)
        self.upload_worker = FeishuUploadWorker(
            app_id=appid, app_secret=appsecret, mode=mode,
            app_token=apptoken, table_id=tableid, record_id=record_id,
            script_field=scriptfield, script_text=script_text,
            folder_token=foldertoken, topic_name=topic_name,
        )

        def on_done(msg):
            self.pbar.setVisible(False)
            self.lbl_status.setText("飞书同步完成。")
            self.show_info(msg)

        def on_err(err_msg):
            self.pbar.setVisible(False)
            self.lbl_status.setText("飞书同步失败。")
            QMessageBox.critical(self.parent_widget, "飞书同步失败", err_msg)

        self.upload_worker.finished.connect(on_done)
        self.upload_worker.error.connect(on_err)
        self.upload_worker.start()
