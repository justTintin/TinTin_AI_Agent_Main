# -*- coding: utf-8 -*-
"""自动上架数据包：导入、结构校验、sku.xlsx 解析。"""
import os
import re
import shutil
import struct
import zipfile
from dataclasses import dataclass, field
from datetime import datetime

from openpyxl import load_workbook

from .config import AUTO_LISTING_SYNC_DIR, DOUYIN_STORES

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}


class ValidationError(RuntimeError):
    pass


@dataclass
class SkuRow:
    name: str
    merchant_code: str = ""


@dataclass
class PackageInfo:
    working_dir: str
    source_name: str
    shop_key: str
    shop_name: str
    title: str = ""
    brand: str = ""
    model: str = ""
    manufacturer: str = ""
    skus: list = field(default_factory=list)
    main_images: list = field(default_factory=list)
    detail_images: list = field(default_factory=list)
    sku_images: list = field(default_factory=list)
    warnings: list = field(default_factory=list)


def normalize_name(name: str) -> str:
    return re.sub(r"[^\u4e00-\u9fa5a-zA-Z0-9]", "", name or "").lower()


def shop_matches(name: str, shop_key: str) -> bool:
    info = DOUYIN_STORES.get(shop_key)
    if not info:
        return False
    cand = normalize_name(name)
    if normalize_name(info.get("name", "")) in cand or cand in normalize_name(info.get("name", "")):
        return True
    return any(normalize_name(a) in cand or cand in normalize_name(a)
               for a in info.get("aliases", []) if a)


def _find_dir(base: str, wanted: str) -> str:
    if not os.path.isdir(base):
        return ""
    direct = os.path.join(base, wanted)
    if os.path.isdir(direct):
        return direct
    for root, dirs, _files in os.walk(base):
        if wanted in dirs:
            return os.path.join(root, wanted)
    return ""


def _collect_images(directory: str) -> list:
    if not directory or not os.path.isdir(directory):
        return []

    def index_of(path):
        base = os.path.splitext(os.path.basename(path))[0]
        m = re.search(r"(\d+)", base)
        return int(m.group(1)) if m else 10 ** 9

    files = [os.path.join(directory, n) for n in os.listdir(directory)
             if os.path.splitext(n)[1].lower() in IMAGE_EXTS]
    files.sort(key=lambda p: (index_of(p), os.path.basename(p).lower()))
    return files


def image_size(path: str):
    """读取 PNG/JPG 尺寸，无需 Pillow；失败返回 None。"""
    try:
        with open(path, "rb") as f:
            data = f.read(32)
        if data[:8] == b"\x89PNG\r\n\x1a\n":
            w, h = struct.unpack(">LL", data[16:24])
            return int(w), int(h)
        if data[:2] == b"\xff\xd8":
            pos = 2
            while pos < len(data):
                if data[pos] != 0xFF:
                    pos += 1
                    continue
                marker = data[pos + 1]
                if marker in (0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9,
                              0xCA, 0xCB, 0xCD, 0xCE, 0xCF):
                    h, w = struct.unpack(">HH", data[pos + 5:pos + 9])
                    return int(w), int(h)
                if marker == 0xD8 or 0xD0 <= marker <= 0xD9:
                    pos += 2
                else:
                    length = struct.unpack(">H", data[pos + 2:pos + 4])[0]
                    pos += 2 + length
    except Exception:
        pass
    return None


def _read_excel(working_dir: str) -> dict:
    xls_path = os.path.join(working_dir, "sku.xlsx")
    if not os.path.isfile(xls_path):
        raise ValidationError(f"数据包缺少 sku.xlsx: {working_dir}")
    try:
        wb = load_workbook(xls_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValidationError(f"读取 sku.xlsx 失败: {e}") from e

    if not wb.worksheets:
        raise ValidationError("sku.xlsx 没有任何工作表")

    sheet1 = wb.worksheets[0]
    headers = [str(c.value or "").strip() if c.value is not None else "" for c in sheet1[1]]
    rows = []
    for values in sheet1.iter_rows(min_row=2, values_only=True):
        rows.append(values)

    def col(name_candidates):
        for cand in name_candidates:
            for i, h in enumerate(headers):
                if h and (h.lower() == cand.lower() or cand.lower() in h.lower()):
                    return i
        return -1

    sku_col = col(["sku图片名", "组合装名称", "SKU"])
    brand_col = col(["品牌"])
    merchant_col = col(["修改后的商品编码", "同步后的商家编码", "商家编码"])
    title_col = col(["商品标题", "标题"])
    model_col = col(["型号"])
    manufacturer_col = col(["生产厂家", "生产厂商"])

    brand = ""
    for values in rows:
        if brand_col >= 0 and len(values) > brand_col and values[brand_col] is not None:
            brand = str(values[brand_col]).strip()
            if brand and brand.lower() != "nan":
                break

    skus = []
    if sku_col >= 0:
        for values in rows:
            if len(values) <= sku_col or values[sku_col] is None:
                continue
            name = " ".join(str(values[sku_col]).strip().split())
            if not name or name.lower() == "nan":
                continue
            code = ""
            if merchant_col >= 0 and len(values) > merchant_col and values[merchant_col] is not None:
                code = str(values[merchant_col]).strip()
            if name not in [s.name for s in skus]:
                skus.append(SkuRow(name=name, merchant_code=code))

    sheet2 = wb.worksheets[1] if len(wb.worksheets) > 1 else None
    title = ""
    kv = {}
    if sheet2 is not None:
        rows2 = list(sheet2.iter_rows(values_only=True))
        if rows2:
            h1 = str(rows2[0][0] or "").strip() if len(rows2[0]) > 0 else ""
            h2 = str(rows2[0][1] or "").strip() if len(rows2[0]) > 1 else ""
            if h2 and not h2.lower().startswith("unnamed"):
                title = h2
            elif len(rows2) > 1 and len(rows2[1]) > 1 and rows2[1][1] is not None:
                title = str(rows2[1][1]).strip()
            for values in rows2[1:]:
                if len(values) >= 2 and values[0] is not None and values[1] is not None:
                    key = str(values[0]).strip()
                    val = str(values[1]).strip()
                    if key and val and val.lower() != "nan":
                        kv[key] = val

    if not title and title_col >= 0:
        for values in rows:
            if len(values) > title_col and values[title_col] is not None:
                title = str(values[title_col]).strip()
                if title:
                    break

    model = kv.get("型号", "")
    manufacturer = kv.get("生产厂家") or kv.get("生产厂商", "")
    if not model and model_col >= 0:
        for values in rows:
            if len(values) > model_col and values[model_col] is not None:
                model = str(values[model_col]).strip()
                break
    if not manufacturer and manufacturer_col >= 0:
        for values in rows:
            if len(values) > manufacturer_col and values[manufacturer_col] is not None:
                manufacturer = str(values[manufacturer_col]).strip()
                break
    return {"title": title, "brand": brand, "model": model,
            "manufacturer": manufacturer, "skus": skus}


def _locate_working_dir(base: str) -> str:
    direct = os.path.join(base, "sku.xlsx")
    if os.path.isfile(direct):
        return base
    for root, _dirs, files in os.walk(base):
        if "sku.xlsx" in files:
            return root
    return base


def inspect_package(working_dir: str, source_name: str, shop_key: str) -> PackageInfo:
    working_dir = _locate_working_dir(working_dir)
    if not os.path.isfile(os.path.join(working_dir, "sku.xlsx")):
        raise ValidationError(f"数据包中未找到 sku.xlsx: {working_dir}")
    if not shop_matches(source_name, shop_key):
        info = DOUYIN_STORES.get(shop_key, {})
        expected = " / ".join([info.get("name", "")] + list(info.get("aliases", [])))
        raise ValidationError(
            f"数据包名称“{source_name}”未包含目标店铺关键词，请包含：{expected}")

    main_dir = _find_dir(working_dir, "主图")
    detail_dir = _find_dir(working_dir, "详情页") or _find_dir(working_dir, "详情")
    sku_dir = _find_dir(working_dir, "sku图")
    main_images = _collect_images(main_dir)
    detail_images = _collect_images(detail_dir)
    sku_images = _collect_images(sku_dir)

    missing = []
    if not main_images:
        missing.append("主图")
    if not detail_images:
        missing.append("详情页")
    if not sku_images:
        missing.append("sku图")
    if missing:
        raise ValidationError(f"数据包缺少可上传图片目录：{'、'.join(missing)}")

    non_square = []
    for p in main_images:
        size = image_size(p)
        if size and size[0] != size[1]:
            non_square.append(os.path.basename(p))
    if non_square:
        raise ValidationError(f"主图必须为 1:1，以下文件不是正方形：{', '.join(non_square[:5])}")

    excel = _read_excel(working_dir)
    skus = excel["skus"]
    if not skus:
        raise ValidationError("sku.xlsx 未解析到任何 SKU/规格行（需要 sku图片名 或 组合装名称 列）")

    warnings = []
    if not excel["title"]:
        warnings.append("未解析到商品标题，发布时可能跳过标题填写")
    if not excel["brand"]:
        excel["brand"] = "无品牌"

    info_data = DOUYIN_STORES.get(shop_key, {})
    return PackageInfo(
        working_dir=working_dir,
        source_name=source_name,
        shop_key=shop_key,
        shop_name=info_data.get("name", shop_key),
        title=excel["title"],
        brand=excel["brand"],
        model=excel["model"],
        manufacturer=excel["manufacturer"],
        skus=skus,
        main_images=main_images,
        detail_images=detail_images,
        sku_images=sku_images,
        warnings=warnings,
    )


def prepare_package(input_path: str, shop_key: str, run_id: str = "") -> PackageInfo:
    if not input_path or not os.path.exists(input_path):
        raise ValidationError(f"输入路径不存在: {input_path}")
    source_name = os.path.basename(input_path.rstrip("/\\"))
    run_id = run_id or datetime.now().strftime("%Y%m%d_%H%M%S")
    batch_name = os.path.splitext(source_name)[0] if source_name.lower().endswith(".zip") else source_name
    batch_root = os.path.join(AUTO_LISTING_SYNC_DIR, batch_name)
    staged_root = os.path.join(batch_root, "_runs", run_id, "input")
    os.makedirs(staged_root, exist_ok=True)

    if os.path.isdir(input_path):
        shutil.copytree(input_path, staged_root, dirs_exist_ok=True)
    elif input_path.lower().endswith(".zip"):
        try:
            with zipfile.ZipFile(input_path, "r") as zf:
                zf.extractall(staged_root)
        except Exception as e:
            raise ValidationError(f"解压数据包失败: {e}") from e
    else:
        raise ValidationError("输入必须是文件夹或 .zip 压缩包")

    return inspect_package(staged_root, source_name, shop_key)
